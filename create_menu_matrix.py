"""
Hotel PMS — 메뉴 구조도 & 권한 매트릭스 Excel 생성 스크립트
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ─── 공통 스타일 정의 ────────────────────────────────────────
def border(style='thin'):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def font(bold=False, color='000000', size=10):
    return Font(bold=bold, color=color, size=size, name='맑은 고딕')

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

COLORS = {
    'header_main':  '1E3A5F',  # 진한 남색
    'header_sub':   '2E6DA4',  # 파란색
    'group_fd':     'D6E4F0',  # 연파랑
    'group_crm':    'D5F5E3',  # 연초록
    'group_ops':    'FEF9E7',  # 연노랑
    'group_set':    'F5EEF8',  # 연보라
    'group_main':   'EBF5FB',  # 아주 연한 파랑
    'full':         '27AE60',  # 초록  ✅
    'view':         '2980B9',  # 파랑  👁
    'edit':         'F39C12',  # 주황  ✏️
    'none':         'E74C3C',  # 빨강  ⛔
    'row_odd':      'FAFAFA',
    'row_even':     'F0F4F8',
    'role_bg':      'EBF5FB',
    'white':        'FFFFFF',
}

ROLE_COLORS = {
    'ADMIN':   '1E3A5F',
    'MANAGER': '154360',
    'FRONT':   '1A5276',
    'HOUSE':   '1B6B35',
    'ANC':     '784212',
    'ACCT':    '6C3483',
}

# ─── 시트 1: 메뉴 구조도 ─────────────────────────────────────
ws1 = wb.active
ws1.title = '메뉴 구조도'
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = 'A3'

headers = ['그룹', '상위 메뉴', '하위 메뉴', '파일 경로', '주요 기능', 'PC', '모바일', '비고']
col_widths = [14, 16, 18, 42, 45, 7, 9, 20]

# 제목 행
ws1.merge_cells('A1:H1')
c = ws1['A1']
c.value = '🏨 Hotel PMS — 메뉴 구조도'
c.font = Font(bold=True, size=14, color='FFFFFF', name='맑은 고딕')
c.fill = fill(COLORS['header_main'])
c.alignment = center()
ws1.row_dimensions[1].height = 28

# 헤더 행
for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    c = ws1.cell(row=2, column=col, value=h)
    c.font = font(bold=True, color='FFFFFF', size=10)
    c.fill = fill(COLORS['header_sub'])
    c.alignment = center()
    c.border = border()
    ws1.column_dimensions[get_column_letter(col)].width = w
ws1.row_dimensions[2].height = 22

# 데이터 정의: (그룹, 상위, 하위, 파일경로, 기능, pc, mobile, 비고, 그룹색)
pages = [
    ('메인', '로그인', '', 'dashboard/login.html', '직원 로그인', '✅', '✅', '', 'group_main'),
    ('메인', '대시보드', '', 'dashboard/dashboard.html', 'KPI 요약, 예약 현황, 하우스키핑 상태', '✅', '✅', '', 'group_main'),
    ('Front Desk', '예약 타임라인', '', 'frontdesk/reservation-timeline.html', '객실별 예약 타임라인, 신규 예약 등록', '✅', '✅', '', 'group_fd'),
    ('Front Desk', '예약 목록', '', 'frontdesk/reservation-list.html', '예약 검색/필터, 상태 관리', '✅', '✅', 'Bottom Sheet 필터', 'group_fd'),
    ('Front Desk', '체크인/아웃', '', 'frontdesk/checkin.html', '체크인·체크아웃 처리, 객실 배정', '✅', '✅', '', 'group_fd'),
    ('Guest & CRM', '투숙객 관리', '', 'crm/guests.html', '고객 프로파일, 투숙 이력, 선호도', '✅', '✅', '', 'group_crm'),
    ('Guest & CRM', 'VIP 멤버십', '', 'crm/membership.html', '멤버 등급 관리, 포인트 현황', '✅', '✅', '', 'group_crm'),
    ('Guest & CRM', 'VIP 멤버십', '↳ 등급 이력', 'crm/tier-history.html', '멤버 등급 변경 이력 조회', '✅', '✅', '멤버십 하위', 'group_crm'),
    ('Operations', '객실 관리', '↳ 객실 현황/목록', 'operations/rooms.html', '객실 상태 현황판, 객실 목록', '✅', '✅', '', 'group_ops'),
    ('Operations', '객실 관리', '↳ 객실/유형 등록', 'operations/room-setup.html', '객실 타입 등록, 정보 수정', '✅', '✅', '', 'group_ops'),
    ('Operations', '요금 캘린더', '', 'operations/rates.html', '날짜별 요금 설정', '✅', '⛔', 'PC 전용', 'group_ops'),
    ('Operations', '하우스키핑', '', 'operations/housekeeping.html', '청소 현황, 상태 변경, 배정', '✅', '✅', '', 'group_ops'),
    ('Operations', '통합 정산', '↳ 정산 목록', 'operations/folio.html', 'Folio 조회, 기간별 매출 요약 (금일/주별/월별/년도별)', '✅', '✅', '', 'group_ops'),
    ('Operations', '통합 정산', '↳ 매출 분석', 'operations/folio-chart.html', '매출 차트, 기간별 비교 (주별/월별/년도별)', '✅', '✅', '', 'group_ops'),
    ('Operations', '부가서비스', '↳ 룸서비스', 'operations/room-service.html', '룸서비스 주문 접수·처리', '✅', '✅', '', 'group_ops'),
    ('Operations', '부가서비스', '↳ 골프장', 'operations/golf.html', '골프장 업체 등록, 예약 및 완료 관리', '✅', '✅', '', 'group_ops'),
    ('Operations', '부가서비스', '↳ 렌트카', 'operations/rentacar.html', '렌트카 업체 등록, 예약 및 완료 관리', '✅', '✅', '', 'group_ops'),
    ('Settings', '호텔 설정', '', 'settings/settings.html', '호텔 기본 정보, 시스템 환경 설정', '✅', '✅', '', 'group_set'),
    ('Settings', '직원 관리', '', 'settings/staff.html', '직원 계정 등록·수정, 권한 설정', '✅', '✅', '', 'group_set'),
]

for i, (grp, parent, child, path, func, pc, mob, note, grp_color) in enumerate(pages):
    row = i + 3
    row_fill = fill(COLORS[grp_color])
    values = [grp, parent, child, path, func, pc, mob, note]
    for col, val in enumerate(values, 1):
        c = ws1.cell(row=row, column=col, value=val)
        c.fill = row_fill
        c.border = border()
        c.font = font(size=9)
        if col in (6, 7):
            c.alignment = center()
        else:
            c.alignment = left()
    ws1.row_dimensions[row].height = 18

# ─── 시트 2: 권한 매트릭스 ───────────────────────────────────
ws2 = wb.create_sheet('권한 매트릭스')
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = 'A4'

roles = ['ADMIN\n시스템관리자', 'MANAGER\n호텔관리자', 'FRONT\n프론트데스크', 'HOUSE\n하우스키핑', 'ANC\n부가서비스', 'ACCT\n회계/정산']
role_keys = ['ADMIN', 'MANAGER', 'FRONT', 'HOUSE', 'ANC', 'ACCT']

# 제목
ws2.merge_cells('A1:I1')
c = ws2['A1']
c.value = '🔐 Hotel PMS — 역할별 권한 매트릭스'
c.font = Font(bold=True, size=14, color='FFFFFF', name='맑은 고딕')
c.fill = fill(COLORS['header_main'])
c.alignment = center()
ws2.row_dimensions[1].height = 28

# 범례
ws2.merge_cells('A2:I2')
c = ws2['A2']
c.value = '범례:  ✅ 전체 접근    👁 조회만    ✏️ 편집 가능    ⛔ 접근 불가'
c.font = Font(bold=True, size=9, color='333333', name='맑은 고딕')
c.fill = fill('F8F9FA')
c.alignment = left()
ws2.row_dimensions[2].height = 18

# 헤더 행
perm_headers = ['그룹', '메뉴명', '파일 경로'] + roles
perm_widths =  [14,     20,       38,          14, 14, 14, 14, 14, 14]
for col, (h, w) in enumerate(zip(perm_headers, perm_widths), 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.font = Font(bold=True, color='FFFFFF', size=9, name='맑은 고딕')
    c.fill = fill(COLORS['header_main'])
    c.alignment = center()
    c.border = border()
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.row_dimensions[3].height = 36

# 권한 데이터: (그룹, 메뉴명, 파일, ADMIN, MANAGER, FRONT, HOUSE, ANC, ACCT, 그룹색)
perm_data = [
    # 메인
    ('메인', '로그인', 'login.html', '✅', '✅', '✅', '✅', '✅', '✅', 'group_main'),
    ('메인', '대시보드', 'dashboard.html', '✅', '✅', '👁', '👁', '👁', '👁', 'group_main'),
    # Front Desk
    ('Front Desk', '예약 타임라인', 'reservation-timeline.html', '✅', '✅', '✏️', '👁', '⛔', '⛔', 'group_fd'),
    ('Front Desk', '예약 목록', 'reservation-list.html', '✅', '✅', '✏️', '⛔', '⛔', '👁', 'group_fd'),
    ('Front Desk', '체크인/아웃', 'checkin.html', '✅', '✅', '✏️', '⛔', '⛔', '⛔', 'group_fd'),
    # Guest & CRM
    ('Guest & CRM', '투숙객 관리', 'guests.html', '✅', '✅', '✏️', '⛔', '⛔', '👁', 'group_crm'),
    ('Guest & CRM', 'VIP 멤버십', 'membership.html', '✅', '✅', '👁', '⛔', '⛔', '⛔', 'group_crm'),
    ('Guest & CRM', '↳ 등급 이력', 'tier-history.html', '✅', '✅', '👁', '⛔', '⛔', '⛔', 'group_crm'),
    # Operations - 객실
    ('Operations', '↳ 객실 현황/목록', 'rooms.html', '✅', '✅', '👁', '👁', '⛔', '⛔', 'group_ops'),
    ('Operations', '↳ 객실/유형 등록', 'room-setup.html', '✅', '✅', '⛔', '⛔', '⛔', '⛔', 'group_ops'),
    ('Operations', '요금 캘린더', 'rates.html', '✅', '✅', '⛔', '⛔', '⛔', '👁', 'group_ops'),
    ('Operations', '하우스키핑', 'housekeeping.html', '✅', '✅', '👁', '✏️', '⛔', '⛔', 'group_ops'),
    # Operations - 정산
    ('Operations', '↳ 정산 목록', 'folio.html', '✅', '✅', '✏️', '⛔', '⛔', '✅', 'group_ops'),
    ('Operations', '↳ 매출 분석', 'folio-chart.html', '✅', '✅', '⛔', '⛔', '⛔', '✅', 'group_ops'),
    # Operations - 부가서비스
    ('Operations', '↳ 룸서비스', 'room-service.html', '✅', '✅', '👁', '⛔', '✏️', '👁', 'group_ops'),
    ('Operations', '↳ 골프장', 'golf.html', '✅', '✅', '👁', '⛔', '✏️', '👁', 'group_ops'),
    ('Operations', '↳ 렌트카', 'rentacar.html', '✅', '✅', '👁', '⛔', '✏️', '👁', 'group_ops'),
    # Settings
    ('Settings', '호텔 설정', 'settings.html', '✅', '👁', '⛔', '⛔', '⛔', '⛔', 'group_set'),
    ('Settings', '직원 관리', 'staff.html', '✅', '✏️', '⛔', '⛔', '⛔', '⛔', 'group_set'),
]

PERM_FILL = {
    '✅':  'D5F5E3',  # 연초록
    '👁':  'D6EAF8',  # 연파랑
    '✏️': 'FEF9E7',  # 연노랑
    '⛔':  'FADBD8',  # 연빨강
}

for i, row_data in enumerate(perm_data):
    row = i + 4
    grp, menu, path = row_data[0], row_data[1], row_data[2]
    perms = row_data[3:9]
    grp_color = row_data[9]
    base_fill = fill(COLORS[grp_color])

    for col, val in enumerate([grp, menu, path], 1):
        c = ws2.cell(row=row, column=col, value=val)
        c.fill = base_fill
        c.border = border()
        c.font = font(size=9)
        c.alignment = left()

    for col, perm in enumerate(perms, 4):
        c = ws2.cell(row=row, column=col, value=perm)
        c.fill = fill(PERM_FILL.get(perm, 'FFFFFF'))
        c.border = border()
        c.font = Font(bold=True, size=11, name='맑은 고딕')
        c.alignment = center()

    ws2.row_dimensions[row].height = 20

# ─── 시트 3: 역할 요약 ───────────────────────────────────────
ws3 = wb.create_sheet('역할 정의')
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = 'A3'

ws3.merge_cells('A1:E1')
c = ws3['A1']
c.value = '👤 역할(Role) 정의 및 접근 범위 요약'
c.font = Font(bold=True, size=14, color='FFFFFF', name='맑은 고딕')
c.fill = fill(COLORS['header_main'])
c.alignment = center()
ws3.row_dimensions[1].height = 28

role_headers = ['역할 코드', '역할명', '설명', '접근 페이지 수', '주요 접근 범위']
role_widths  = [14,         16,       50,     16,               50]
for col, (h, w) in enumerate(zip(role_headers, role_widths), 1):
    c = ws3.cell(row=2, column=col, value=h)
    c.font = font(bold=True, color='FFFFFF')
    c.fill = fill(COLORS['header_sub'])
    c.alignment = center()
    c.border = border()
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.row_dimensions[2].height = 22

role_summary = [
    ('ADMIN',   '시스템 관리자', '모든 메뉴·기능 전체 접근, 시스템 설정 포함',     '19 / 19', '전체'),
    ('MANAGER', '호텔 관리자',   '통계·매출·직원 관리 포함, 설정 일부 접근',       '18 / 19', '설정 조회, 매출 통계, 직원 편집'),
    ('FRONT',   '프론트 데스크', '예약·체크인/아웃·고객·정산 목록 접근',           '12 / 19', '예약·체크인·정산 목록·고객'),
    ('ACCT',    '회계/정산',     '통합 정산·매출 분석·다운로드 전용 접근',          '9 / 19',  '정산 목록·매출 분석·부가서비스 조회'),
    ('ANC',     '부가서비스 직원','룸서비스·골프장·렌트카 접수·처리 접근',          '4 / 19',  '대시보드·룸서비스·골프장·렌트카'),
    ('HOUSE',   '하우스키핑',    '하우스키핑·객실 현황 조회만 접근',               '3 / 19',  '대시보드·객실 현황·하우스키핑'),
]

role_row_colors = ['D6EAF8', 'D5F5E3', 'FEF9E7', 'F5EEF8', 'FDEBD0', 'FADBD8']

for i, (code, name, desc, cnt, scope) in enumerate(role_summary):
    row = i + 3
    vals = [code, name, desc, cnt, scope]
    rc = role_row_colors[i]
    for col, val in enumerate(vals, 1):
        c = ws3.cell(row=row, column=col, value=val)
        c.fill = fill(rc)
        c.border = border()
        c.font = Font(bold=(col == 1), size=9, name='맑은 고딕',
                      color=ROLE_COLORS.get(code, '000000') if col == 1 else '000000')
        c.alignment = center() if col in (1, 4) else left()
    ws3.row_dimensions[row].height = 22

# ─── 저장 ────────────────────────────────────────────────────
out_path = r'e:\AI_Project\Hotel_PMS\Hotel_PMS_메뉴구조_권한매트릭스.xlsx'
wb.save(out_path)
print(f'[OK] 저장 완료: {out_path}')
