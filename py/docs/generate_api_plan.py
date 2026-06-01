import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

data_schemas = {
    "Dashboard API": [
        ("GET", "/api/v1/dashboard/kpis", "Response", "occupancyRate", "Number", "Y", "객실 점유율 (%)", "85.4"),
        ("GET", "/api/v1/dashboard/kpis", "Response", "adr", "Number", "Y", "평균 일일 객실 단가 (ADR)", "145.50"),
        ("GET", "/api/v1/dashboard/kpis", "Response", "revpar", "Number", "Y", "판매 가능 객실당 수익 (RevPAR)", "124.25"),
        ("GET", "/api/v1/dashboard/kpis", "Response", "totalRevenue", "Number", "Y", "총 매출", "24500.00"),
        
        ("GET", "/api/v1/dashboard/revenue-chart", "Response", "date", "String", "Y", "차트 X축 날짜", "2026-05-27"),
        ("GET", "/api/v1/dashboard/revenue-chart", "Response", "roomRev", "Number", "Y", "객실 매출", "18500"),
        ("GET", "/api/v1/dashboard/revenue-chart", "Response", "fnbRev", "Number", "Y", "F&B (식음료) 매출", "3200"),
        ("GET", "/api/v1/dashboard/revenue-chart", "Response", "spaRev", "Number", "Y", "스파 매출", "1400"),
        ("GET", "/api/v1/dashboard/revenue-chart", "Response", "otherRev", "Number", "Y", "기타 매출", "1400"),
        
        ("GET", "/api/v1/dashboard/recent-tasks", "Response", "taskId", "String", "Y", "작업 ID", "TSK-001"),
        ("GET", "/api/v1/dashboard/recent-tasks", "Response", "roomNumber", "String", "Y", "관련 객실 번호", "1401"),
        ("GET", "/api/v1/dashboard/recent-tasks", "Response", "taskType", "String", "Y", "작업 분류", "cleaning, maintenance"),
        ("GET", "/api/v1/dashboard/recent-tasks", "Response", "status", "String", "Y", "진행 상태", "pending, progress, done"),
        ("GET", "/api/v1/dashboard/recent-tasks", "Response", "createdAt", "String", "Y", "작업 등록 일시", "2026-05-27T10:30:00Z"),
    ],
    "FrontDesk API": [
        ("GET", "/api/v1/frontdesk/checkins", "Response", "reservationId", "String", "Y", "예약 고유 ID", "RES-1001"),
        ("GET", "/api/v1/frontdesk/checkins", "Response", "guestName", "String", "Y", "투숙객 이름", "홍길동"),
        ("GET", "/api/v1/frontdesk/checkins", "Response", "guestTier", "String", "N", "투숙객 멤버십 등급", "VIP, Regular 등"),
        ("GET", "/api/v1/frontdesk/checkins", "Response", "checkInDate", "String", "Y", "체크인 날짜", "2026-05-27"),
        ("GET", "/api/v1/frontdesk/checkins", "Response", "checkOutDate", "String", "Y", "체크아웃 날짜", "2026-05-29"),
        ("GET", "/api/v1/frontdesk/checkins", "Response", "roomTypeCode", "String", "Y", "객실 타입 코드", "DELUXE_DBL"),
        ("GET", "/api/v1/frontdesk/checkins", "Response", "roomNumber", "String", "N", "배정된 객실 번호", "1401"),
        ("GET", "/api/v1/frontdesk/checkins", "Response", "reservationStatus", "String", "Y", "예약 상태", "confirmed, in-house"),
        ("GET", "/api/v1/frontdesk/checkins", "Response", "balanceAmount", "Number", "Y", "잔여 결제 금액", "350.00"),
        
        ("GET", "/api/v1/frontdesk/stayovers", "Response", "roomNumber", "String", "Y", "객실 번호", "1401"),
        ("GET", "/api/v1/frontdesk/stayovers", "Response", "frontStatus", "String", "Y", "객실 상태", "occupied, vacant"),
        ("GET", "/api/v1/frontdesk/stayovers", "Response", "guestFlag", "String", "N", "고객 요청 상태", "dnd, mur, away, none"),
        
        ("POST", "/api/v1/frontdesk/rooms/guest-flag", "Request", "roomNumber", "String", "Y", "대상 객실 번호", "1401"),
        ("POST", "/api/v1/frontdesk/rooms/guest-flag", "Request", "flag", "String", "Y", "변경할 상태 값", "dnd, mur, away, none"),
    ],
    "Reservation API": [
        ("GET", "/api/v1/reservations", "Response", "reservationId", "String", "Y", "예약 고유 ID", "RES-1001"),
        ("GET", "/api/v1/reservations", "Response", "guestName", "String", "Y", "투숙객 이름", "홍길동"),
        ("GET", "/api/v1/reservations", "Response", "guestPhone", "String", "N", "투숙객 연락처", "010-1234-5678"),
        ("GET", "/api/v1/reservations", "Response", "checkInDate", "String", "Y", "체크인 날짜", "2026-05-27"),
        ("GET", "/api/v1/reservations", "Response", "checkOutDate", "String", "Y", "체크아웃 날짜", "2026-05-29"),
        ("GET", "/api/v1/reservations", "Response", "roomTypeCode", "String", "Y", "객실 타입 코드", "DELUXE_DBL"),
        ("GET", "/api/v1/reservations", "Response", "roomNumber", "String", "N", "배정된 객실 번호", "1401"),
        ("GET", "/api/v1/reservations", "Response", "reservationStatus", "String", "Y", "예약 상태", "confirmed, cancelled, no-show"),
        ("GET", "/api/v1/reservations", "Response", "bookingSource", "String", "Y", "유입 채널", "OTA, Direct"),
        ("GET", "/api/v1/reservations", "Response", "totalAmount", "Number", "Y", "총 결제액", "500.00"),
        ("GET", "/api/v1/reservations", "Response", "paidAmount", "Number", "Y", "기결제액", "100.00"),
        ("GET", "/api/v1/reservations", "Response", "balanceAmount", "Number", "Y", "잔여 결제액", "400.00"),
        
        ("GET", "/api/v1/reservations/{reservationId}", "Response", "reservationId", "String", "Y", "상세 조회용 예약ID", "RES-1001"),
        ("GET", "/api/v1/reservations/{reservationId}", "Response", "guestId", "String", "N", "고객 고유 번호", "GST-001"),
        ("GET", "/api/v1/reservations/{reservationId}", "Response", "guestEmail", "String", "N", "고객 이메일", "hong@example.com"),
        ("GET", "/api/v1/reservations/{reservationId}", "Response", "paxAdults", "Number", "Y", "성인 인원수", "2"),
        ("GET", "/api/v1/reservations/{reservationId}", "Response", "paxChildren", "Number", "Y", "아동 인원수", "0"),
        ("GET", "/api/v1/reservations/{reservationId}", "Response", "specialRequests", "String", "N", "특별 요청사항", "고층 객실 희망"),
    ],
    "Groups API": [
        ("GET", "/api/v1/groups/blocks", "Response", "blockId", "String", "Y", "단체 고유 ID", "GRP-001"),
        ("GET", "/api/v1/groups/blocks", "Response", "groupName", "String", "Y", "단체명 / 행사명", "글로벌 IT 서밋"),
        ("GET", "/api/v1/groups/blocks", "Response", "companyName", "String", "N", "연결된 기업 고객사", "삼성전자"),
        ("GET", "/api/v1/groups/blocks", "Response", "startDate", "String", "Y", "블록 시작 날짜", "2026-06-10"),
        ("GET", "/api/v1/groups/blocks", "Response", "endDate", "String", "Y", "블록 종료 날짜", "2026-06-15"),
        ("GET", "/api/v1/groups/blocks", "Response", "blockStatus", "String", "Y", "블록 상태", "tentative, definite, cancelled"),
        ("GET", "/api/v1/groups/blocks", "Response", "roomTypeCode", "String", "Y", "할당된 객실 타입", "DELUXE_DBL"),
        ("GET", "/api/v1/groups/blocks", "Response", "totalBlockedRooms", "Number", "Y", "총 할당(블록) 객실 수", "50"),
        ("GET", "/api/v1/groups/blocks", "Response", "pickedUpRooms", "Number", "Y", "실제 예약된(픽업) 객실 수", "35"),
        ("GET", "/api/v1/groups/blocks", "Response", "agreedRate", "Number", "Y", "협의된 특가(Rate)", "120.00"),
        ("GET", "/api/v1/groups/blocks", "Response", "cutoffDate", "String", "Y", "미사용 객실 반환 기준일", "2026-05-20"),
    ],
    "Operations API": [
        ("GET", "/api/v1/operations/housekeeping/rooms", "Response", "roomNumber", "String", "Y", "객실 번호", "1401"),
        ("GET", "/api/v1/operations/housekeeping/rooms", "Response", "roomTypeCode", "String", "Y", "객실 타입 코드", "DELUXE_DBL"),
        ("GET", "/api/v1/operations/housekeeping/rooms", "Response", "frontStatus", "String", "Y", "프론트 기준 상태", "vacant, occupied, oos"),
        ("GET", "/api/v1/operations/housekeeping/rooms", "Response", "hkStatus", "String", "Y", "청소 상태", "clean, dirty, inspect"),
        ("GET", "/api/v1/operations/housekeeping/rooms", "Response", "guestFlag", "String", "N", "고객 요청 상태", "dnd, mur, away"),
        ("GET", "/api/v1/operations/housekeeping/rooms", "Response", "activeTaskId", "String", "N", "진행중인 작업 ID", "TSK-101"),
        
        ("GET", "/api/v1/operations/housekeeping/tasks", "Response", "taskId", "String", "Y", "작업 ID", "TSK-101"),
        ("GET", "/api/v1/operations/housekeeping/tasks", "Response", "roomNumber", "String", "Y", "대상 객실 번호", "1401"),
        ("GET", "/api/v1/operations/housekeeping/tasks", "Response", "taskType", "String", "Y", "작업 종류", "cleaning, inspection"),
        ("GET", "/api/v1/operations/housekeeping/tasks", "Response", "taskStatus", "String", "Y", "진행 상태", "pending, progress, done"),
        ("GET", "/api/v1/operations/housekeeping/tasks", "Response", "assigneeName", "String", "N", "배정된 담당자 이름", "Maria G"),
        
        ("GET", "/api/v1/operations/pos/orders", "Response", "orderId", "String", "Y", "주문 고유 ID", "ORD-5001"),
        ("GET", "/api/v1/operations/pos/orders", "Response", "roomNumber", "String", "Y", "비용을 청구할 객실 번호", "1401"),
        ("GET", "/api/v1/operations/pos/orders", "Response", "category", "String", "Y", "부가서비스 카테고리", "roomservice, minibar, spa, laundry, other"),
        ("GET", "/api/v1/operations/pos/orders", "Response", "guestName", "String", "N", "투숙객 이름", "Steve T"),
        ("GET", "/api/v1/operations/pos/orders", "Response", "orderSummary", "String", "Y", "주문 요약 문자열", "클럽 샌드위치, 콜라"),
        ("GET", "/api/v1/operations/pos/orders", "Response", "totalAmount", "Number", "Y", "총 주문 금액", "32.00"),
        ("GET", "/api/v1/operations/pos/orders", "Response", "orderStatus", "String", "Y", "주문 처리 상태", "new, prep, done"),
        ("GET", "/api/v1/operations/pos/orders", "Response", "orderedAt", "String", "Y", "주문 접수 시간", "2026-05-27T10:05:00Z"),
    ],
    "CRM API": [
        ("GET", "/api/v1/crm/guests", "Response", "guestId", "String", "Y", "고객 프로필 고유 ID", "GST-001"),
        ("GET", "/api/v1/crm/guests", "Response", "guestName", "String", "Y", "고객 성함", "김철수"),
        ("GET", "/api/v1/crm/guests", "Response", "phoneNumber", "String", "N", "고객 연락처", "010-1234-5678"),
        ("GET", "/api/v1/crm/guests", "Response", "email", "String", "N", "고객 이메일", "chulsoo@example.com"),
        ("GET", "/api/v1/crm/guests", "Response", "tier", "String", "Y", "고객 멤버십 등급", "Bronze, Silver, Gold, Platinum"),
        ("GET", "/api/v1/crm/guests", "Response", "totalStays", "Number", "Y", "역대 총 투숙 횟수", "5"),
        ("GET", "/api/v1/crm/guests", "Response", "totalSpend", "Number", "Y", "역대 총 결제액", "1500.00"),
        ("GET", "/api/v1/crm/guests", "Response", "lastVisitDate", "String", "N", "마지막 투숙일", "2026-03-15"),
    ],
    "Settings API": [
        ("GET", "/api/v1/settings/users", "Response", "userId", "String", "Y", "계정 고유 ID", "USR-01"),
        ("GET", "/api/v1/settings/users", "Response", "username", "String", "Y", "로그인 아이디", "admin_01"),
        ("GET", "/api/v1/settings/users", "Response", "fullName", "String", "Y", "직원 성함", "김매니저"),
        ("GET", "/api/v1/settings/users", "Response", "roleName", "String", "Y", "할당된 권한 그룹", "Admin, Front Desk"),
        ("GET", "/api/v1/settings/users", "Response", "department", "String", "Y", "소속 부서", "프론트 오피스"),
        ("GET", "/api/v1/settings/users", "Response", "isActive", "Boolean", "Y", "계정 활성화 상태", "true, false"),
        
        ("GET", "/api/v1/settings/roles", "Response", "roleId", "String", "Y", "역할 고유 ID", "ROL-01"),
        ("GET", "/api/v1/settings/roles", "Response", "roleName", "String", "Y", "역할 이름", "프론트 데스크 직원"),
        ("GET", "/api/v1/settings/roles", "Response", "permissions", "Array[String]", "Y", "허용된 권한코드 배열", "['view_reservations', 'edit_reservations']"),
    ]
}

columns = ["Method", "API Endpoint", "Req/Res", "Field Name", "Data Type", "Mandatory (Y/N)", "Field Description", "Sample Value"]

file_path = r'E:\AI_Project\Hotel_PMS\docs\API_Integration_Plan.xlsx'

with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    for sheet_name, rows in data_schemas.items():
        df = pd.DataFrame(rows, columns=columns)
        # sheet names can't exceed 31 characters
        safe_sheet_name = sheet_name[:31]
        df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
        
        # Formatting
        workbook = writer.book
        worksheet = writer.sheets[safe_sheet_name]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
        border = Border(left=Side(style='thin', color='BFBFBF'), 
                        right=Side(style='thin', color='BFBFBF'), 
                        top=Side(style='thin', color='BFBFBF'), 
                        bottom=Side(style='thin', color='BFBFBF'))
        
        for col_num, value in enumerate(df.columns.values):
            cell = worksheet.cell(row=1, column=col_num+1)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=len(columns)):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                
        # Auto-adjust column widths
        for col in worksheet.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    val_str = str(cell.value)
                    length = sum(2 if ord(c) > 127 else 1 for c in val_str) # Korean text logic
                    if length > max_length:
                        max_length = length
                except:
                    pass
            adjusted_width = min(max_length + 2, 45)
            if column_letter == 'B': # Endpoint
                adjusted_width = 40
            elif column_letter == 'D': # Field Name
                adjusted_width = 25
            elif column_letter == 'G': # Description column
                adjusted_width = 40
            elif column_letter == 'H': # Sample data column
                adjusted_width = 30
            worksheet.column_dimensions[column_letter].width = adjusted_width

print(f"API Integration Plan Excel file successfully generated at: {file_path}")
