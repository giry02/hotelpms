import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

api_data = [
    {
        "domain": "Dashboard_API",
        "apis": [
            {
                "name": "핵심 지표 (KPI) 조회",
                "method": "GET",
                "uri": "/api/v1/dashboard/kpis",
                "desc": "선택한 날짜의 KPI 데이터를 조회합니다.",
                "request": [
                    ("targetDate", "String (Query)", "Y", "조회 기준 일자 (YYYY-MM-DD)", "2026-05-27")
                ],
                "response": [
                    ("occupancyRate", "Number", "Y", "객실 점유율 (%)", "85.4"),
                    ("adr", "Number", "Y", "평균 일일 객실 단가", "145.5"),
                    ("revpar", "Number", "Y", "객실당 수익", "124.25"),
                    ("totalRevenue", "Number", "Y", "총 매출", "24500.00")
                ]
            },
            {
                "name": "매출 추이 차트 조회",
                "method": "GET",
                "uri": "/api/v1/dashboard/revenue-chart",
                "desc": "주간/월간 등 특정 기간의 매출 추이를 조회합니다.",
                "request": [
                    ("startDate", "String (Query)", "Y", "조회 시작일 (YYYY-MM-DD)", "2026-05-20"),
                    ("endDate", "String (Query)", "Y", "조회 종료일 (YYYY-MM-DD)", "2026-05-27")
                ],
                "response": [
                    ("chartData", "Array[Object]", "Y", "일자별 매출 데이터 배열", "[...]"),
                    ("chartData[].date", "String", "Y", "기준 날짜", "2026-05-20"),
                    ("chartData[].roomRev", "Number", "Y", "객실 매출", "18500"),
                    ("chartData[].fnbRev", "Number", "Y", "식음료 매출", "3200"),
                    ("chartData[].spaRev", "Number", "Y", "스파 매출", "1400"),
                    ("chartData[].otherRev", "Number", "Y", "기타 매출", "500")
                ]
            },
            {
                "name": "최근 작업 요약 조회",
                "method": "GET",
                "uri": "/api/v1/dashboard/recent-tasks",
                "desc": "대시보드 위젯에 표시할 최근 하우스키핑/보수 작업을 조회합니다.",
                "request": [
                    ("limit", "Number (Query)", "N", "조회 건수 (기본 5)", "5")
                ],
                "response": [
                    ("tasks", "Array[Object]", "Y", "작업 목록 배열", "[...]"),
                    ("tasks[].taskId", "String", "Y", "작업 ID", "TSK-001"),
                    ("tasks[].roomNumber", "String", "Y", "관련 객실 번호", "1401"),
                    ("tasks[].taskType", "String", "Y", "작업 분류 (cleaning, maintenance)", "cleaning"),
                    ("tasks[].status", "String", "Y", "진행 상태 (pending, progress, done)", "pending"),
                    ("tasks[].createdAt", "String", "Y", "등록 일시", "2026-05-27T10:30:00Z")
                ]
            }
        ]
    },
    {
        "domain": "Reservation_API",
        "apis": [
            {
                "name": "당일 체크인/아웃 리스트 조회",
                "method": "GET",
                "uri": "/api/v1/reservations/checkins",
                "desc": "프론트데스크에서 오늘 체크인/체크아웃 할 예약 목록을 조회합니다.",
                "request": [
                    ("targetDate", "String (Query)", "Y", "기준 날짜 (YYYY-MM-DD)", "2026-05-27"),
                    ("type", "String (Query)", "Y", "목록 타입 (arrival, departure, stayover)", "arrival"),
                    ("searchKeyword", "String (Query)", "N", "투숙객 이름 또는 예약번호 검색", "홍길동"),
                    ("page", "Number (Query)", "N", "페이지 번호 (기본 1)", "1"),
                    ("limit", "Number (Query)", "N", "페이지당 건수 (기본 20)", "20")
                ],
                "response": [
                    ("totalCount", "Number", "Y", "총 검색 건수", "15"),
                    ("items", "Array[Object]", "Y", "예약 목록 배열", "[...]"),
                    ("items[].reservationId", "String", "Y", "예약 고유 ID", "RES-1001"),
                    ("items[].guestName", "String", "Y", "투숙객 이름", "홍길동"),
                    ("items[].guestTier", "String", "N", "멤버십 등급 (VIP 등)", "VIP"),
                    ("items[].checkInDate", "String", "Y", "체크인 날짜", "2026-05-27"),
                    ("items[].checkOutDate", "String", "Y", "체크아웃 날짜", "2026-05-29"),
                    ("items[].roomTypeCode", "String", "Y", "객실 타입", "DELUXE_DBL"),
                    ("items[].roomNumber", "String", "N", "배정된 객실 (없을 수 있음)", "1401"),
                    ("items[].status", "String", "Y", "예약 상태 (confirmed, in-house)", "confirmed"),
                    ("items[].balanceAmount", "Number", "Y", "잔여 결제 금액", "350.00")
                ]
            },
            {
                "name": "전체 예약 목록 조회 (예약관리 그리드)",
                "method": "GET",
                "uri": "/api/v1/reservations",
                "desc": "조건에 맞는 전체 예약 목록을 페이징하여 조회합니다.",
                "request": [
                    ("startDate", "String (Query)", "N", "체크인 기간 검색 시작일", "2026-05-01"),
                    ("endDate", "String (Query)", "N", "체크인 기간 검색 종료일", "2026-05-31"),
                    ("status", "String (Query)", "N", "상태 필터링", "confirmed"),
                    ("keyword", "String (Query)", "N", "이름/연락처/번호 검색", "010-1234-5678"),
                    ("page", "Number (Query)", "Y", "요청 페이지 번호", "1"),
                    ("size", "Number (Query)", "Y", "페이지당 항목 수", "50")
                ],
                "response": [
                    ("totalElements", "Number", "Y", "전체 데이터 수", "120"),
                    ("totalPages", "Number", "Y", "전체 페이지 수", "3"),
                    ("content", "Array[Object]", "Y", "예약 목록", "[...]"),
                    ("content[].reservationId", "String", "Y", "예약 ID", "RES-1001"),
                    ("content[].guestName", "String", "Y", "투숙객명", "김철수"),
                    ("content[].guestPhone", "String", "N", "연락처", "010-1111-2222"),
                    ("content[].checkInDate", "String", "Y", "체크인일", "2026-05-10"),
                    ("content[].checkOutDate", "String", "Y", "체크아웃일", "2026-05-12"),
                    ("content[].roomTypeCode", "String", "Y", "객실 타입", "STANDARD"),
                    ("content[].roomNumber", "String", "N", "객실 번호", "1002"),
                    ("content[].status", "String", "Y", "예약 상태", "confirmed"),
                    ("content[].source", "String", "Y", "예약 채널 (OTA, Direct)", "OTA"),
                    ("content[].totalAmount", "Number", "Y", "총 결제 예정액", "250.00"),
                    ("content[].paidAmount", "Number", "Y", "기결제액", "50.00"),
                    ("content[].balanceAmount", "Number", "Y", "잔여 결제액", "200.00")
                ]
            },
            {
                "name": "예약 상세 정보 조회 (팝업용)",
                "method": "GET",
                "uri": "/api/v1/reservations/{reservationId}",
                "desc": "예약건에 대한 모든 상세 정보를 불러옵니다. (상세 수정 모달 오픈 시)",
                "request": [
                    ("reservationId", "String (Path)", "Y", "조회할 예약 ID", "RES-1001")
                ],
                "response": [
                    ("reservationId", "String", "Y", "예약 ID", "RES-1001"),
                    ("guestId", "String", "N", "CRM 고객 ID", "GST-801"),
                    ("guestName", "String", "Y", "투숙객명", "김철수"),
                    ("guestEmail", "String", "N", "이메일", "kim@example.com"),
                    ("guestPhone", "String", "Y", "연락처", "010-1111-2222"),
                    ("checkInDate", "String", "Y", "체크인", "2026-05-10"),
                    ("checkOutDate", "String", "Y", "체크아웃", "2026-05-12"),
                    ("roomTypeCode", "String", "Y", "객실타입", "STANDARD"),
                    ("roomNumber", "String", "N", "배정객실", "1002"),
                    ("status", "String", "Y", "예약상태", "confirmed"),
                    ("paxAdults", "Number", "Y", "성인 수", "2"),
                    ("paxChildren", "Number", "Y", "아동 수", "0"),
                    ("specialRequests", "String", "N", "특별요청사항", "고층 객실 배정 요망"),
                    ("totalAmount", "Number", "Y", "총액", "250.00"),
                    ("payments", "Array[Object]", "N", "결제 내역 배열", "[{id:'PAY-1', amount:50.0}]")
                ]
            }
        ]
    },
    {
        "domain": "Operations_API",
        "apis": [
            {
                "name": "하우스키핑 객실 상태 그리드 조회",
                "method": "GET",
                "uri": "/api/v1/operations/rooms/status",
                "desc": "하우스키핑 화면에서 전체 객실의 현재 청소 및 투숙 상태를 조회합니다.",
                "request": [
                    ("floor", "String (Query)", "N", "특정 층 필터링", "14"),
                    ("hkStatus", "String (Query)", "N", "특정 청소 상태 필터링 (clean/dirty/inspect)", "dirty")
                ],
                "response": [
                    ("rooms", "Array[Object]", "Y", "객실 목록 배열", "[...]"),
                    ("rooms[].roomNumber", "String", "Y", "객실 번호", "1401"),
                    ("rooms[].roomTypeCode", "String", "Y", "객실 타입", "DELUXE_DBL"),
                    ("rooms[].frontStatus", "String", "Y", "프론트 점유 상태 (vacant, occupied, oos)", "occupied"),
                    ("rooms[].hkStatus", "String", "Y", "청소 상태 (clean, dirty, inspect)", "dirty"),
                    ("rooms[].guestFlag", "String", "N", "고객 요청 (dnd, mur, away)", "dnd"),
                    ("rooms[].activeTaskId", "String", "N", "진행중인 작업 ID (있을경우)", "TSK-101")
                ]
            },
            {
                "name": "고객 요청 상태 변경 (DND/MUR)",
                "method": "POST",
                "uri": "/api/v1/operations/rooms/{roomNumber}/guest-flag",
                "desc": "특정 객실의 고객 요청 플래그를 변경합니다.",
                "request": [
                    ("roomNumber", "String (Path)", "Y", "대상 객실 번호", "1401"),
                    ("flag", "String (Body JSON)", "Y", "변경할 상태 (dnd, mur, away, none)", "mur")
                ],
                "response": [
                    ("success", "Boolean", "Y", "처리 성공 여부", "true"),
                    ("updatedFlag", "String", "Y", "적용된 최종 상태", "mur")
                ]
            },
            {
                "name": "통합 POS 주문 내역 조회",
                "method": "GET",
                "uri": "/api/v1/operations/pos/orders",
                "desc": "룸서비스, 미니바, 스파 등의 부가서비스 청구 내역을 조회합니다.",
                "request": [
                    ("date", "String (Query)", "N", "조회 일자", "2026-05-27"),
                    ("category", "String (Query)", "N", "주문 카테고리 필터 (roomservice, minibar, spa, laundry)", "roomservice"),
                    ("status", "String (Query)", "N", "상태 (new, prep, done)", "new")
                ],
                "response": [
                    ("orders", "Array[Object]", "Y", "주문 내역 배열", "[...]"),
                    ("orders[].orderId", "String", "Y", "주문 ID", "ORD-5001"),
                    ("orders[].roomNumber", "String", "Y", "청구 객실", "1401"),
                    ("orders[].category", "String", "Y", "서비스 분류", "roomservice"),
                    ("orders[].guestName", "String", "N", "투숙객 이름", "Steve T"),
                    ("orders[].orderSummary", "String", "Y", "항목 요약 문자열", "클럽 샌드위치 외 1건"),
                    ("orders[].totalAmount", "Number", "Y", "청구 총액", "32.00"),
                    ("orders[].orderStatus", "String", "Y", "상태 (new, prep, done)", "new"),
                    ("orders[].orderedAt", "String", "Y", "주문 일시", "2026-05-27T10:05:00Z")
                ]
            }
        ]
    }
]

# Create Excel
wb = Workbook()
wb.remove(wb.active) # remove default sheet

# styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
sub_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
sub_header_font = Font(bold=True)
border = Border(left=Side(style='thin', color='A6A6A6'), 
                right=Side(style='thin', color='A6A6A6'), 
                top=Side(style='thin', color='A6A6A6'), 
                bottom=Side(style='thin', color='A6A6A6'))
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

for domain_data in api_data:
    ws = wb.create_sheet(title=domain_data['domain'])
    
    # Set columns widths
    ws.column_dimensions['A'].width = 15 # 구분 (I/O)
    ws.column_dimensions['B'].width = 30 # 필드명
    ws.column_dimensions['C'].width = 20 # 타입
    ws.column_dimensions['D'].width = 12 # 필수
    ws.column_dimensions['E'].width = 45 # 설명
    ws.column_dimensions['F'].width = 30 # 샘플
    
    row_idx = 1
    for api in domain_data['apis']:
        # API Title Block
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        cell = ws.cell(row=row_idx, column=1, value=f"■ [{api['method']}] {api['name']}")
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row_idx].height = 25
        row_idx += 1
        
        # API Meta Info
        ws.cell(row=row_idx, column=1, value="엔드포인트(URI)").font = sub_header_font
        ws.cell(row=row_idx, column=1).fill = sub_header_fill
        ws.cell(row=row_idx, column=1).border = border
        ws.cell(row=row_idx, column=1).alignment = align_center
        
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
        c = ws.cell(row=row_idx, column=2, value=api['uri'])
        c.border = border
        c.font = Font(bold=True, color="C00000") # Dark red for URI
        # add borders for merged cells manually
        for col in range(2, 7): ws.cell(row=row_idx, column=col).border = border
        row_idx += 1
        
        ws.cell(row=row_idx, column=1, value="기능 설명").font = sub_header_font
        ws.cell(row=row_idx, column=1).fill = sub_header_fill
        ws.cell(row=row_idx, column=1).border = border
        ws.cell(row=row_idx, column=1).alignment = align_center
        
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
        c = ws.cell(row=row_idx, column=2, value=api['desc'])
        c.border = border
        for col in range(2, 7): ws.cell(row=row_idx, column=col).border = border
        row_idx += 1
        
        # Request Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        req_title = ws.cell(row=row_idx, column=1, value="▼ Request (Input)")
        req_title.font = header_font
        req_title.fill = header_fill
        req_title.alignment = align_center
        row_idx += 1
        
        headers = ["구분 (I/O)", "파라미터 / 필드명", "데이터 타입", "필수 (M/O)", "설명 (Description)", "샘플 (Sample)"]
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=h)
            c.font = sub_header_font
            c.fill = sub_header_fill
            c.border = border
            c.alignment = align_center
        row_idx += 1
        
        if not api['request']:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
            c = ws.cell(row=row_idx, column=1, value="요청 파라미터 없음 (No Input Parameters)")
            c.alignment = align_center
            c.border = border
            for col in range(1, 7): ws.cell(row=row_idx, column=col).border = border
            row_idx += 1
        else:
            for req in api['request']:
                row_data = ["Request", req[0], req[1], req[2], req[3], req[4]]
                for col_idx, val in enumerate(row_data, 1):
                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.border = border
                    c.alignment = align_left if col_idx in [2, 5] else align_center
                row_idx += 1
                
        # Response Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        res_title = ws.cell(row=row_idx, column=1, value="▼ Response (Output)")
        res_title.font = header_font
        res_title.fill = header_fill
        res_title.alignment = align_center
        row_idx += 1
        
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=h)
            c.font = sub_header_font
            c.fill = sub_header_fill
            c.border = border
            c.alignment = align_center
        row_idx += 1
        
        for res in api['response']:
            row_data = ["Response", res[0], res[1], res[2], res[3], res[4]]
            for col_idx, val in enumerate(row_data, 1):
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.border = border
                c.alignment = align_left if col_idx in [2, 5] else align_center
            row_idx += 1
            
        row_idx += 3 # gap between APIs

file_path = r'E:\AI_Project\Hotel_PMS\docs\API_Integration_Plan_v3.xlsx'
wb.save(file_path)
print(f"Professional API Specification generated at: {file_path}")
