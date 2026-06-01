import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

api_data = [
    {
        "domain": "1_Dashboard",
        "apis": [
            {
                "name": "대시보드 KPI 조회", "method": "GET", "uri": "/api/v1/dashboard/kpis", "desc": "일일 점유율, ADR, RevPAR, 총매출 조회",
                "request": [("date", "Query", "Y", "기준 날짜", "2026-05-27")],
                "response": [("occupancyRate", "Number", "Y", "점유율", "85.4"), ("adr", "Number", "Y", "객실단가", "145.5"), ("revpar", "Number", "Y", "객실당수익", "124.2"), ("revenue", "Number", "Y", "매출", "24500")]
            },
            {
                "name": "매출 추이 차트", "method": "GET", "uri": "/api/v1/dashboard/revenue-chart", "desc": "날짜별 매출 데이터 배열 (차트용)",
                "request": [("startDate", "Query", "Y", "시작일", "2026-05-20"), ("endDate", "Query", "Y", "종료일", "2026-05-27")],
                "response": [("items", "Array", "Y", "데이터 배열", "[]"), ("items[].date", "String", "Y", "날짜", "2026-05-20"), ("items[].roomRev", "Number", "Y", "객실매출", "18500"), ("items[].fnbRev", "Number", "Y", "F&B매출", "3200")]
            }
        ]
    },
    {
        "domain": "2_Frontdesk",
        "apis": [
            {
                "name": "체크인/아웃 대상 목록 조회", "method": "GET", "uri": "/api/v1/frontdesk/checkins", "desc": "당일 체크인, 체크아웃 리스트 화면 전용",
                "request": [("date", "Query", "Y", "날짜", "2026-05-27"), ("type", "Query", "Y", "타입 (arrival/departure)", "arrival"), ("keyword", "Query", "N", "이름/예약번호 검색", "홍길동")],
                "response": [
                    ("items", "Array", "Y", "목록", "[]"), 
                    ("items[].reservationId", "String", "Y", "예약ID", "RES-1001"), 
                    ("items[].guestName", "String", "Y", "투숙객명", "홍길동"),
                    ("items[].vipTier", "String", "N", "VIP/멤버십 등급 (UI 아이콘용)", "Gold"),
                    ("items[].roomTypeCode", "String", "Y", "객실타입", "DLX"), 
                    ("items[].roomNumber", "String", "N", "배정호실", "1401"), 
                    ("items[].checkInDate", "String", "Y", "체크인", "2026-05-27"),
                    ("items[].checkOutDate", "String", "Y", "체크아웃", "2026-05-29"),
                    ("items[].balance", "Number", "Y", "잔액", "350"),
                    ("items[].status", "String", "Y", "현재상태", "confirmed")
                ]
            },
            {
                "name": "전체 예약 목록 조회 (그리드)", "method": "GET", "uri": "/api/v1/reservations", "desc": "예약 리스트 페이지 전용 (고객 상세정보 배제, 리스트 출력용 최적화)",
                "request": [
                    ("startDate", "Query", "N", "검색 시작일", "2026-05-01"), 
                    ("endDate", "Query", "N", "종료일", "2026-05-31"), 
                    ("status", "Query", "N", "상태 필터", "confirmed"),
                    ("page", "Query", "Y", "페이지 번호", "1"),
                    ("size", "Query", "Y", "페이지 크기", "20")
                ],
                "response": [
                    ("totalElements", "Number", "Y", "검색된 전체수", "120"), 
                    ("items", "Array", "Y", "배열", "[]"), 
                    ("items[].reservationId", "String", "Y", "예약번호", "RES-1001"),
                    ("items[].guestName", "String", "Y", "투숙객명", "Alexander"),
                    ("items[].vipTier", "String", "N", "VIP 정보", "Platinum"),
                    ("items[].roomNumber", "String", "N", "호실", "401"),
                    ("items[].roomTypeCode", "String", "Y", "객실유형", "Deluxe"),
                    ("items[].checkInDate", "String", "Y", "체크인", "2026-05-12"),
                    ("items[].checkOutDate", "String", "Y", "체크아웃", "2026-05-15"),
                    ("items[].nights", "Number", "Y", "숙박일수", "3"),
                    ("items[].bookingSource", "String", "Y", "유입채널", "Direct"),
                    ("items[].isB2B", "Boolean", "Y", "기업/B2B 여부", "false"),
                    ("items[].totalAmount", "Number", "Y", "총금액(USD)", "450"),
                    ("items[].status", "String", "Y", "상태", "checkedin")
                ]
            },
            {
                "name": "신규 예약 등록", "method": "POST", "uri": "/api/v1/reservations", "desc": "신규 예약 생성",
                "request": [("guestName", "Body", "Y", "이름", "이순신"), ("guestPhone", "Body", "Y", "연락처", "010-9999-8888"), ("checkInDate", "Body", "Y", "체크인", "2026-06-01"), ("checkOutDate", "Body", "Y", "체크아웃", "2026-06-05"), ("roomTypeCode", "Body", "Y", "객실타입", "DLX"), ("adults", "Body", "Y", "성인수", "2")],
                "response": [("success", "Boolean", "Y", "성공여부", "true"), ("reservationId", "String", "Y", "생성된 예약ID", "RES-1005")]
            },
            {
                "name": "예약 상태 변경", "method": "PUT", "uri": "/api/v1/reservations/{id}/status", "desc": "체크인/아웃/취소 처리",
                "request": [("id", "Path", "Y", "예약ID", "RES-1001"), ("status", "Body", "Y", "변경할 상태", "checked-in")],
                "response": [("success", "Boolean", "Y", "성공", "true")]
            }
        ]
    },
    {
        "domain": "3_Operations",
        "apis": [
            {
                "name": "객실 상태 조회", "method": "GET", "uri": "/api/v1/operations/housekeeping/rooms", "desc": "하우스키핑용 객실 현황 (동/층/상태)",
                "request": [("building", "Query", "N", "건물", "Main Tower"), ("floor", "Query", "N", "층수", "14")],
                "response": [
                    ("items", "Array", "Y", "배열", "[]"), 
                    ("items[].roomNumber", "String", "Y", "호실", "1401"), 
                    ("items[].bldg", "String", "Y", "건물동", "Main Tower"),
                    ("items[].floor", "String", "Y", "층수", "14F"),
                    ("items[].roomTypeCode", "String", "Y", "타입", "Deluxe"),
                    ("items[].hkStatus", "String", "Y", "청소상태", "dirty"), 
                    ("items[].frontStatus", "String", "Y", "프론트상태", "occupied")
                ]
            },
            {
                "name": "운영 태스크 리스트", "method": "GET", "uri": "/api/v1/operations/tasks", "desc": "대시보드 및 하우스키핑의 작업 요청 내역",
                "request": [("status", "Query", "N", "상태필터", "dirty")],
                "response": [
                    ("items", "Array", "Y", "목록", "[]"),
                    ("items[].taskId", "String", "Y", "작업ID", "t1"),
                    ("items[].roomNumber", "String", "Y", "대상호실", "1401"),
                    ("items[].taskType", "String", "Y", "분류(checkout/stayover/deep)", "checkout"),
                    ("items[].status", "String", "Y", "현재상태", "dirty"),
                    ("items[].priority", "Boolean", "Y", "긴급여부", "true"),
                    ("items[].note", "String", "N", "요청메모", "VIP 체크인 예정 (14:00)")
                ]
            },
            {
                "name": "청소/유지보수 작업 등록", "method": "POST", "uri": "/api/v1/operations/tasks", "desc": "새로운 작업 요청",
                "request": [("roomNumber", "Body", "Y", "호실", "1401"), ("taskType", "Body", "Y", "분류", "cleaning"), ("note", "Body", "N", "메모", "수건 추가")],
                "response": [("success", "Boolean", "Y", "성공", "true"), ("taskId", "String", "Y", "작업ID", "t200")]
            },
            {
                "name": "작업 상태 업데이트", "method": "PUT", "uri": "/api/v1/operations/tasks/{id}/status", "desc": "작업 완료 등 상태 변경",
                "request": [("id", "Path", "Y", "작업ID", "t1"), ("status", "Body", "Y", "상태", "clean")],
                "response": [("success", "Boolean", "Y", "성공", "true")]
            }
        ]
    },
    {
        "domain": "4_CRM",
        "apis": [
            {
                "name": "고객(Guest) 리스트 조회", "method": "GET", "uri": "/api/v1/crm/guests", "desc": "CRM 고객 목록 (데이터베이스에서 집계된 정보 포함)",
                "request": [("keyword", "Query", "N", "검색", "Alexander")],
                "response": [
                    ("items", "Array", "Y", "배열", "[]"), 
                    ("items[].guestId", "String", "Y", "고객ID", "G-001"),
                    ("items[].guestName", "String", "Y", "이름", "Alexander"),
                    ("items[].totalStays", "Number", "Y", "총방문수", "5"),
                    ("items[].lastVisitDate", "String", "N", "최근방문일", "2026-02-10"),
                    ("items[].totalSpend", "Number", "Y", "총결제액", "4500"),
                    ("items[].cancelCount", "Number", "Y", "취소건수", "0"),
                    ("items[].phone", "String", "N", "연락처", "+1 555-0102"),
                    ("items[].email", "String", "N", "이메일", "alex@example.com"),
                    ("items[].country", "String", "N", "국적", "United States"),
                    ("items[].tier", "String", "Y", "멤버십", "Platinum")
                ]
            }
        ]
    }
]

# Create Excel
wb = Workbook()
wb.remove(wb.active)

header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
sub_header_fill = PatternFill(start_color="EAEFF7", end_color="EAEFF7", fill_type="solid")
sub_header_font = Font(bold=True)
border = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'), top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

for domain_data in api_data:
    ws = wb.create_sheet(title=domain_data['domain'])
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 25
    
    row_idx = 1
    for api in domain_data['apis']:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        cell = ws.cell(row=row_idx, column=1, value=f"■ [{api['method']}] {api['name']}")
        cell.font = Font(bold=True, size=11, color="1F4E78")
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1
        
        ws.cell(row=row_idx, column=1, value="URI").font = sub_header_font
        ws.cell(row=row_idx, column=1).fill = sub_header_fill
        ws.cell(row=row_idx, column=1).border = border
        ws.cell(row=row_idx, column=1).alignment = align_center
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
        c = ws.cell(row=row_idx, column=2, value=api['uri'])
        c.border = border
        c.font = Font(bold=True, color="C00000")
        for col in range(2, 7): ws.cell(row=row_idx, column=col).border = border
        row_idx += 1
        
        ws.cell(row=row_idx, column=1, value="설명").font = sub_header_font
        ws.cell(row=row_idx, column=1).fill = sub_header_fill
        ws.cell(row=row_idx, column=1).border = border
        ws.cell(row=row_idx, column=1).alignment = align_center
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
        c = ws.cell(row=row_idx, column=2, value=api['desc'])
        c.border = border
        for col in range(2, 7): ws.cell(row=row_idx, column=col).border = border
        row_idx += 1
        
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        req_title = ws.cell(row=row_idx, column=1, value="▶ Request (Input)")
        req_title.font = header_font
        req_title.fill = header_fill
        req_title.alignment = align_center
        row_idx += 1
        
        headers = ["구분 (I/O)", "파라미터/바디 필드", "데이터 타입", "필수 (Y/N)", "설명 (Description)", "샘플 (Sample)"]
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=h)
            c.font = sub_header_font
            c.fill = sub_header_fill
            c.border = border
            c.alignment = align_center
        row_idx += 1
        
        if not api['request']:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
            c = ws.cell(row=row_idx, column=1, value="요청 파라미터 없음")
            c.alignment = align_center
            c.border = border
            for col in range(1, 7): ws.cell(row=row_idx, column=col).border = border
            row_idx += 1
        else:
            for req in api['request']:
                row_data = ["Request", req[0], req[1], req[2], req[3], req[4]]
                for col_idx, val in enumerate(row_data, 1):
                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.border = border
                    c.alignment = align_left if col_idx in [2, 5] else align_center
                row_idx += 1
                
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        res_title = ws.cell(row=row_idx, column=1, value="▶ Response (Output)")
        res_title.font = header_font
        res_title.fill = header_fill
        res_title.alignment = align_center
        row_idx += 1
        
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=h)
            c.font = sub_header_font
            c.fill = sub_header_fill
            c.border = border
            c.alignment = align_center
        row_idx += 1
        
        for res in api['response']:
            row_data = ["Response", res[0], res[1], res[2], res[3], res[4]]
            for col_idx, val in enumerate(row_data, 1):
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.border = border
                c.alignment = align_left if col_idx in [2, 5] else align_center
            row_idx += 1
            
        row_idx += 2 

file_path = r'E:\AI_Project\Hotel_PMS\docs\API_Integration_Plan_Perfect_v6.xlsx'
wb.save(file_path)
print(f"Perfect Overfetch-free API Specification generated at: {file_path}")
