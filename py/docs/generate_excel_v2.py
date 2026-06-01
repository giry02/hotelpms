import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define detailed schema for each page in Korean
data_schemas = {
    "dashboard.html": [
        ("대시보드 요약", "getKPIs()", "GET", "date", "String", "Y", "KPI 기준 날짜", "2026-05-27"),
        ("대시보드 요약", "getKPIs()", "GET", "occupancy", "Number", "Y", "객실 점유율 (%)", "85.4"),
        ("대시보드 요약", "getKPIs()", "GET", "adr", "Number", "Y", "평균 일일 객실 단가 (ADR)", "145.50"),
        ("대시보드 요약", "getKPIs()", "GET", "revpar", "Number", "Y", "판매 가능 객실당 수익 (RevPAR)", "124.25"),
        ("대시보드 요약", "getKPIs()", "GET", "revenue", "Number", "Y", "총 매출", "24500.00"),
        
        ("매출 차트", "getRevenueChart()", "GET", "date", "String", "Y", "차트 X축 날짜", "2026-05-27"),
        ("매출 차트", "getRevenueChart()", "GET", "rooms", "Number", "Y", "객실 매출", "18500"),
        ("매출 차트", "getRevenueChart()", "GET", "fnb", "Number", "Y", "F&B (식음료) 매출", "3200"),
        ("매출 차트", "getRevenueChart()", "GET", "spa", "Number", "Y", "스파 매출", "1400"),
        ("매출 차트", "getRevenueChart()", "GET", "other", "Number", "Y", "기타 매출", "1400"),
        
        ("최근 작업", "getTasks()", "GET", "id", "String", "Y", "작업 ID", "TSK-001"),
        ("최근 작업", "getTasks()", "GET", "room", "String", "Y", "관련 객실 번호", "1401"),
        ("최근 작업", "getTasks()", "GET", "type", "String", "Y", "작업 분류", "cleaning(청소), maintenance(보수)"),
        ("최근 작업", "getTasks()", "GET", "status", "String", "Y", "진행 상태", "pending(대기), progress(진행중), done(완료)"),
        ("최근 작업", "getTasks()", "GET", "time", "String", "Y", "작업 등록 시간", "10:30"),
        ("최근 작업", "getTasks()", "GET", "assignee", "String", "N", "담당자 이름", "Maria G"),
    ],
    "checkin.html": [
        ("체크인/아웃 리스트", "getReservations()", "GET", "id", "String", "Y", "예약 고유 ID", "RES-1001"),
        ("체크인/아웃 리스트", "getReservations()", "GET", "guest.name", "String", "Y", "투숙객 이름", "홍길동"),
        ("체크인/아웃 리스트", "getReservations()", "GET", "guest.phone", "String", "N", "투숙객 연락처", "010-1234-5678"),
        ("체크인/아웃 리스트", "getReservations()", "GET", "guest.type", "String", "N", "투숙객 멤버십 등급", "VIP, Regular 등"),
        ("체크인/아웃 리스트", "getReservations()", "GET", "checkIn", "String", "Y", "체크인 날짜 (YYYY-MM-DD)", "2026-05-27"),
        ("체크인/아웃 리스트", "getReservations()", "GET", "checkOut", "String", "Y", "체크아웃 날짜 (YYYY-MM-DD)", "2026-05-29"),
        ("체크인/아웃 리스트", "getReservations()", "GET", "roomType", "String", "Y", "객실 타입 코드", "DELUXE_DBL"),
        ("체크인/아웃 리스트", "getReservations()", "GET", "room", "String", "N", "배정된 객실 번호", "1401"),
        ("체크인/아웃 리스트", "getReservations()", "GET", "status", "String", "Y", "예약 상태", "confirmed(확정), in-house(투숙중), checked-out(체크아웃)"),
        ("체크인/아웃 리스트", "getReservations()", "GET", "balance", "Number", "Y", "잔여 결제 금액", "350.00"),
        ("체크인/아웃 리스트", "getReservations()", "GET", "pax.adults", "Number", "Y", "성인 인원 수", "2"),
        
        ("재실 객실 리스트", "getRooms()", "GET", "id", "String", "Y", "객실 번호", "1401"),
        ("재실 객실 리스트", "getRooms()", "GET", "status", "String", "Y", "객실 점유 상태", "occupied(사용중), vacant-clean(빈방-청소됨)"),
        ("재실 객실 리스트", "getRooms()", "GET", "guestFlag", "String", "N", "고객 요청 상태", "dnd(방해금지), mur(청소요청), away(외출), none(없음)"),
        
        ("고객 요청 상태 변경", "setGuestFlag()", "POST", "roomId", "String", "Y", "대상 객실 번호", "1401"),
        ("고객 요청 상태 변경", "setGuestFlag()", "POST", "flag", "String", "Y", "변경할 상태 값", "dnd, mur, away, none"),
    ],
    "reservation-list.html": [
        ("전체 예약 조회", "getReservations()", "GET", "id", "String", "Y", "예약 고유 ID", "RES-1001"),
        ("전체 예약 조회", "getReservations()", "GET", "guest", "Object", "Y", "투숙객 정보 객체 (이름, 연락처 등)", "{name: '홍길동', phone: '010...'}"),
        ("전체 예약 조회", "getReservations()", "GET", "checkIn", "String", "Y", "체크인 날짜", "2026-05-27"),
        ("전체 예약 조회", "getReservations()", "GET", "checkOut", "String", "Y", "체크아웃 날짜", "2026-05-29"),
        ("전체 예약 조회", "getReservations()", "GET", "roomType", "String", "Y", "객실 타입", "DELUXE_DBL"),
        ("전체 예약 조회", "getReservations()", "GET", "status", "String", "Y", "예약 상태", "confirmed, cancelled(취소), no-show(노쇼)"),
        ("전체 예약 조회", "getReservations()", "GET", "source", "String", "Y", "예약 유입 경로", "OTA(아고다/부킹닷컴 등), Direct(직접예약)"),
        ("전체 예약 조회", "getReservations()", "GET", "amount", "Number", "Y", "총 결제 예정 금액", "500.00"),
        ("전체 예약 조회", "getReservations()", "GET", "paid", "Number", "Y", "기 결제 금액", "100.00"),
        ("전체 예약 조회", "getReservations()", "GET", "balance", "Number", "Y", "잔여 결제 금액", "400.00"),
        
        ("예약 생성/수정", "saveReservation()", "POST/PUT", "reservationData", "Object", "Y", "예약 전체 데이터 객체", "(GET 응답 구조와 동일)"),
    ],
    "groups_blocks.html": [
        ("단체 예약(블록)", "getGroups()", "GET", "id", "String", "Y", "단체 고유 ID", "GRP-001"),
        ("단체 예약(블록)", "getGroups()", "GET", "name", "String", "Y", "단체명 / 행사명", "글로벌 IT 서밋"),
        ("단체 예약(블록)", "getGroups()", "GET", "company", "String", "N", "연결된 기업 고객사", "삼성전자"),
        ("단체 예약(블록)", "getGroups()", "GET", "dates.start", "String", "Y", "블록 시작 날짜", "2026-06-10"),
        ("단체 예약(블록)", "getGroups()", "GET", "dates.end", "String", "Y", "블록 종료 날짜", "2026-06-15"),
        ("단체 예약(블록)", "getGroups()", "GET", "status", "String", "Y", "블록 상태", "tentative(가계약), definite(확정), cancelled(취소)"),
        ("단체 예약(블록)", "getGroups()", "GET", "block.type", "String", "Y", "할당된 객실 타입", "DELUXE_DBL"),
        ("단체 예약(블록)", "getGroups()", "GET", "block.blocked", "Number", "Y", "총 할당(블록) 객실 수", "50"),
        ("단체 예약(블록)", "getGroups()", "GET", "block.picked", "Number", "Y", "실제 예약된(픽업) 객실 수", "35"),
        ("단체 예약(블록)", "getGroups()", "GET", "block.rate", "Number", "Y", "협의된 특가(Rate)", "120.00"),
        ("단체 예약(블록)", "getGroups()", "GET", "cutoff", "String", "Y", "미사용 객실 반환 기준일(Cut-off)", "2026-05-20"),
    ],
    "housekeeping.html": [
        ("하우스키핑 그리드", "getRooms()", "GET", "id", "String", "Y", "객실 번호", "1401"),
        ("하우스키핑 그리드", "getRooms()", "GET", "type", "String", "Y", "객실 타입 코드", "DELUXE_DBL"),
        ("하우스키핑 그리드", "getRooms()", "GET", "status", "String", "Y", "프론트 기준 객실 상태", "vacant(빈방), occupied(사용중), oos(고장)"),
        ("하우스키핑 그리드", "getRooms()", "GET", "hkStatus", "String", "Y", "청소 상태", "clean(청소됨), dirty(청소요망), inspect(점검대기)"),
        ("하우스키핑 그리드", "getRooms()", "GET", "guestFlag", "String", "N", "고객 요청 상태", "dnd(방해금지), mur(청소요청)"),
        ("하우스키핑 그리드", "getRooms()", "GET", "taskId", "String", "N", "연결된 청소 작업 ID (있을 경우)", "TSK-101"),
        
        ("청소 작업 목록", "getTasks()", "GET", "id", "String", "Y", "작업 ID", "TSK-101"),
        ("청소 작업 목록", "getTasks()", "GET", "room", "String", "Y", "대상 객실 번호", "1401"),
        ("청소 작업 목록", "getTasks()", "GET", "type", "String", "Y", "작업 종류", "cleaning(청소), inspection(점검)"),
        ("청소 작업 목록", "getTasks()", "GET", "status", "String", "Y", "진행 상태", "pending, progress, done"),
        ("청소 작업 목록", "getTasks()", "GET", "assignee", "String", "N", "배정된 메이드/담당자 이름", "Maria G"),
    ],
    "unified-pos.html": [
        ("통합 부가서비스", "getOrders()", "GET", "id", "String", "Y", "주문 고유 ID", "ORD-5001"),
        ("통합 부가서비스", "getOrders()", "GET", "room", "String", "Y", "비용을 청구할 객실 번호", "1401"),
        ("통합 부가서비스", "getOrders()", "GET", "type", "String", "Y", "부가서비스 카테고리", "roomservice(룸서비스), minibar(미니바), spa(스파), laundry(세탁), other(기타)"),
        ("통합 부가서비스", "getOrders()", "GET", "guest", "String", "N", "투숙객 이름", "Steve T"),
        ("통합 부가서비스", "getOrders()", "GET", "items", "String", "Y", "주문 항목 내용 (배열 또는 쉼표 구분 문자열)", "클럽 샌드위치, 콜라"),
        ("통합 부가서비스", "getOrders()", "GET", "total", "Number", "Y", "총 주문 금액", "32.00"),
        ("통합 부가서비스", "getOrders()", "GET", "status", "String", "Y", "주문 처리 상태", "new(신규접수), prep(준비중), done(완료)"),
        ("통합 부가서비스", "getOrders()", "GET", "time", "String", "Y", "주문 접수 시간", "10:05"),
        
        ("부가서비스 주문", "saveOrder()", "POST", "orderData", "Object", "Y", "주문 상세 데이터 객체", "(GET 응답 구조와 동일)")
    ],
    "guest-profiles.html": [
        ("고객 CRM", "getGuests()", "GET", "id", "String", "Y", "고객 프로필 고유 ID", "GST-001"),
        ("고객 CRM", "getGuests()", "GET", "name", "String", "Y", "고객 성함", "김철수"),
        ("고객 CRM", "getGuests()", "GET", "phone", "String", "N", "고객 연락처", "010-1234-5678"),
        ("고객 CRM", "getGuests()", "GET", "email", "String", "N", "고객 이메일", "chulsoo@example.com"),
        ("고객 CRM", "getGuests()", "GET", "tier", "String", "Y", "고객 멤버십 등급", "Bronze, Silver, Gold, Platinum"),
        ("고객 CRM", "getGuests()", "GET", "totalStays", "Number", "Y", "역대 총 투숙 횟수", "5"),
        ("고객 CRM", "getGuests()", "GET", "totalSpend", "Number", "Y", "역대 총 결제액", "1500.00"),
        ("고객 CRM", "getGuests()", "GET", "preferences", "String", "N", "고객 특별 요청사항 (선호도)", "고층 객실 선호, 금연실"),
        ("고객 CRM", "getGuests()", "GET", "lastVisit", "String", "N", "마지막 투숙일", "2026-03-15"),
    ],
    "settings.html": [
        ("시스템 계정 관리", "getUsers()", "GET", "id", "String", "Y", "계정 고유 ID", "USR-01"),
        ("시스템 계정 관리", "getUsers()", "GET", "username", "String", "Y", "로그인 아이디", "admin_01"),
        ("시스템 계정 관리", "getUsers()", "GET", "name", "String", "Y", "직원 성함", "김매니저"),
        ("시스템 계정 관리", "getUsers()", "GET", "role", "String", "Y", "할당된 권한 그룹", "admin, manager, frontdesk, housekeep"),
        ("시스템 계정 관리", "getUsers()", "GET", "department", "String", "Y", "소속 부서", "프론트 오피스, 객실 관리부 등"),
        ("시스템 계정 관리", "getUsers()", "GET", "status", "String", "Y", "계정 활성화 상태", "active(활성), suspended(정지)"),
        
        ("권한(역할) 관리", "getRoles()", "GET", "id", "String", "Y", "역할 고유 ID", "ROL-01"),
        ("권한(역할) 관리", "getRoles()", "GET", "name", "String", "Y", "역할 이름", "프론트 데스크 직원"),
        ("권한(역할) 관리", "getRoles()", "GET", "perms", "Array[String]", "Y", "허용된 메뉴/기능 권한 목록 배열", "['view_reservations', 'edit_reservations']"),
    ]
}

columns = ["구분 (Category)", "API 엔드포인트 / 설명", "요청 방식 (Method)", "필드명 (Field Name)", "데이터 타입 (Data Type)", "필수 여부 (Mandatory)", "필드 설명 (Description)", "샘플 데이터 및 허용 값 (Sample / Allowed Values)"]

file_path = r'E:\AI_Project\Hotel_PMS\docs\PMS_API_Specification_Detailed.xlsx'

with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    for sheet_name, rows in data_schemas.items():
        df = pd.DataFrame(rows, columns=columns)
        safe_sheet_name = sheet_name.replace(".html", "")[:31]
        df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
        
        # Formatting
        workbook = writer.book
        worksheet = writer.sheets[safe_sheet_name]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col_num, value in enumerate(df.columns.values):
            cell = worksheet.cell(row=1, column=col_num+1)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=len(columns)):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                
        # Auto-adjust column widths
        for col in worksheet.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        # 한글이 포함되어 있으면 너비를 약간 더 줌
                        val_str = str(cell.value)
                        length = sum(2 if ord(c) > 127 else 1 for c in val_str) # 한글은 2자리 취급
                        if length > max_length:
                            max_length = length
                except:
                    pass
            adjusted_width = min(max_length + 2, 45)
            if column_letter == 'G': # Description column
                adjusted_width = 50
            elif column_letter == 'H': # Sample data column
                adjusted_width = 40
            worksheet.column_dimensions[column_letter].width = adjusted_width

print(f"Korean Detailed Excel file successfully generated at: {file_path}")
