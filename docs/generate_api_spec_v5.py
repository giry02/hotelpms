import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

api_data = [
    {
        "domain": "1_Dashboard",
        "apis": [
            {
                "name": "대시보드 KPI 조회", "method": "GET", "uri": "/api/v1/dashboard/kpis", "desc": "일일 점유율, ADR, RevPAR, 총매출 조회",
                "request": [("date", "Query", "Y", "기준 날짜", "2026-05-27")],
                "response": [("occupancyRate", "Number", "Y", "점유율", "85.4"), ("adr", "Number", "Y", "객실단가", "145.5"), ("revpar", "Number", "Y", "객실당수익", "124.2"), ("revenue", "Number", "Y", "매출", "24500")]
            },
            {
                "name": "매출 추이 차트", "method": "GET", "uri": "/api/v1/dashboard/revenue-chart", "desc": "날짜별 매출 데이터 배열 (차트용)",
                "request": [("startDate", "Query", "Y", "시작일", "2026-05-20"), ("endDate", "Query", "Y", "종료일", "2026-05-27")],
                "response": [("items", "Array", "Y", "데이터 배열", "[]"), ("items[].date", "String", "Y", "날짜", "2026-05-20"), ("items[].roomRev", "Number", "Y", "객실매출", "18500"), ("items[].fnbRev", "Number", "Y", "F&B매출", "3200")]
            }
        ]
    },
    {
        "domain": "2_Frontdesk",
        "apis": [
            {
                "name": "체크인/아웃 대상 조회", "method": "GET", "uri": "/api/v1/frontdesk/checkins", "desc": "당일/지정일 체크인, 체크아웃 리스트",
                "request": [("date", "Query", "Y", "날짜", "2026-05-27"), ("type", "Query", "Y", "타입 (arrival/departure)", "arrival"), ("keyword", "Query", "N", "검색어", "홍길동")],
                "response": [("items", "Array", "Y", "목록", "[]"), ("items[].reservationId", "String", "Y", "예약ID", "RES-1001"), ("items[].guestName", "String", "Y", "투숙객명", "홍길동"), ("items[].roomTypeCode", "String", "Y", "객실타입", "DLX"), ("items[].roomNumber", "String", "N", "호실", "1401"), ("items[].balance", "Number", "Y", "잔액", "350")]
            },
            {
                "name": "전체 예약 목록 조회", "method": "GET", "uri": "/api/v1/reservations", "desc": "전체 예약 목록 (그리드용)",
                "request": [("startDate", "Query", "N", "검색 시작일", "2026-05-01"), ("endDate", "Query", "N", "종료일", "2026-05-31"), ("status", "Query", "N", "상태", "confirmed")],
                "response": [("totalElements", "Number", "Y", "전체수", "120"), ("items", "Array", "Y", "배열", "[]"), ("items[].reservationId", "String", "Y", "예약ID", "RES-1001")]
            },
            {
                "name": "신규 예약 등록", "method": "POST", "uri": "/api/v1/reservations", "desc": "신규 워크인/전화 예약을 시스템에 등록",
                "request": [("guestName", "Body", "Y", "이름", "이순신"), ("guestPhone", "Body", "Y", "연락처", "010-9999-8888"), ("checkInDate", "Body", "Y", "체크인", "2026-06-01"), ("checkOutDate", "Body", "Y", "체크아웃", "2026-06-05"), ("roomTypeCode", "Body", "Y", "객실타입", "DLX"), ("adults", "Body", "Y", "성인수", "2")],
                "response": [("success", "Boolean", "Y", "성공여부", "true"), ("reservationId", "String", "Y", "생성된 예약ID", "RES-1005")]
            },
            {
                "name": "예약 상태 변경 (체크인/아웃 처리)", "method": "PUT", "uri": "/api/v1/reservations/{id}/status", "desc": "특정 예약건의 상태 변경 (체크인 완료, 체크아웃 완료 등)",
                "request": [("id", "Path", "Y", "예약ID", "RES-1001"), ("status", "Body", "Y", "변경할 상태 (in-house, checked-out, cancelled)", "in-house")],
                "response": [("success", "Boolean", "Y", "성공여부", "true"), ("updatedStatus", "String", "Y", "최종상태", "in-house")]
            },
            {
                "name": "예약 타임라인 조회", "method": "GET", "uri": "/api/v1/frontdesk/timeline", "desc": "객실/날짜별 테이프 차트 데이터",
                "request": [("startDate", "Query", "Y", "시작일", "2026-05-27")],
                "response": [("rooms", "Array", "Y", "객실배열", "[]"), ("rooms[].roomId", "String", "Y", "호실", "1401"), ("rooms[].reservations", "Array", "Y", "예약배열", "[...]")]
            },
            {
                "name": "단체 예약(블록) 등록", "method": "POST", "uri": "/api/v1/frontdesk/groups", "desc": "새로운 단체 행사용 블록 생성",
                "request": [("name", "Body", "Y", "단체명", "IT 컨퍼런스"), ("startDate", "Body", "Y", "시작일", "2026-07-01"), ("endDate", "Body", "Y", "종료일", "2026-07-05"), ("blockedRooms", "Body", "Y", "할당 객실수", "30")],
                "response": [("success", "Boolean", "Y", "성공여부", "true"), ("blockId", "String", "Y", "블록ID", "GRP-05")]
            }
        ]
    },
    {
        "domain": "3_Operations",
        "apis": [
            {
                "name": "객실 청소/점유 상태 조회", "method": "GET", "uri": "/api/v1/operations/housekeeping/rooms", "desc": "하우스키핑용 전체 객실 상태 조회",
                "request": [("floor", "Query", "N", "층수", "14")],
                "response": [("items", "Array", "Y", "배열", "[]"), ("items[].roomNumber", "String", "Y", "호실", "1401"), ("items[].hkStatus", "String", "Y", "청소", "dirty"), ("items[].frontStatus", "String", "Y", "프론트상태", "occupied")]
            },
            {
                "name": "청소/유지보수 작업 등록", "method": "POST", "uri": "/api/v1/operations/housekeeping/tasks", "desc": "새로운 작업(청소, 수리 등)을 생성 및 배정",
                "request": [("roomNumber", "Body", "Y", "호실", "1401"), ("taskType", "Body", "Y", "분류(cleaning, maintenance)", "cleaning"), ("assigneeId", "Body", "N", "담당자 계정ID", "USR-05"), ("notes", "Body", "N", "전달사항", "타월 추가 요청")],
                "response": [("success", "Boolean", "Y", "성공", "true"), ("taskId", "String", "Y", "작업ID", "TSK-202")]
            },
            {
                "name": "작업 상태 변경", "method": "PUT", "uri": "/api/v1/operations/housekeeping/tasks/{taskId}/status", "desc": "작업 진행중/완료 처리",
                "request": [("taskId", "Path", "Y", "작업ID", "TSK-202"), ("status", "Body", "Y", "상태 (progress, done)", "done")],
                "response": [("success", "Boolean", "Y", "성공", "true")]
            },
            {
                "name": "객실 고장(OOS) 보고", "method": "POST", "uri": "/api/v1/operations/maintenance", "desc": "해당 객실을 판매 불가 상태로 전환",
                "request": [("roomNumber", "Body", "Y", "호실", "1401"), ("type", "Body", "Y", "분류(oos/ooh)", "oos"), ("description", "Body", "Y", "고장사유", "에어컨 누수")],
                "response": [("success", "Boolean", "Y", "성공", "true"), ("issueId", "String", "Y", "이슈ID", "MNT-05")]
            },
            {
                "name": "POS 부가서비스 주문 등록", "method": "POST", "uri": "/api/v1/operations/pos", "desc": "룸서비스 등 부가서비스 비용을 객실에 청구",
                "request": [("roomNumber", "Body", "Y", "호실", "1401"), ("category", "Body", "Y", "카테고리(roomservice 등)", "roomservice"), ("amount", "Body", "Y", "금액", "45.00"), ("items", "Body", "Y", "주문내역 문자열", "와인 1병, 치즈플래터")],
                "response": [("success", "Boolean", "Y", "성공", "true"), ("orderId", "String", "Y", "주문ID", "ORD-150")]
            },
            {
                "name": "나이트 오딧 실행", "method": "POST", "uri": "/api/v1/operations/night-audit/run", "desc": "일일 마감 실행 (미도착 No-show 처리, 숙박비 일괄 청구, 시스템 날짜 변경)",
                "request": [("confirm", "Body", "Y", "실행 확인 플래그", "true")],
                "response": [("success", "Boolean", "Y", "성공여부", "true"), ("newBusinessDate", "String", "Y", "변경된 신규 영업일", "2026-05-28"), ("message", "String", "Y", "메시지", "마감이 완료되었습니다.")]
            },
            {
                "name": "개별 정산서(Folio) 청구 내역 조회", "method": "GET", "uri": "/api/v1/operations/folio/{resId}", "desc": "객실 결제 및 부가서비스 총 청구서",
                "request": [("resId", "Path", "Y", "예약ID", "RES-1001")],
                "response": [("items", "Array", "Y", "청구내역 배열", "[]"), ("totalAmount", "Number", "Y", "총액", "500"), ("paidAmount", "Number", "Y", "기결제금액", "100"), ("balanceDue", "Number", "Y", "미수금", "400")]
            },
            {
                "name": "결제(Payment) 처리", "method": "POST", "uri": "/api/v1/operations/folio/{resId}/payment", "desc": "미수금에 대한 결제 처리 (카드/현금)",
                "request": [("resId", "Path", "Y", "예약ID", "RES-1001"), ("amount", "Body", "Y", "결제금액", "400"), ("method", "Body", "Y", "결제수단 (credit_card, cash)", "credit_card")],
                "response": [("success", "Boolean", "Y", "성공", "true"), ("paymentId", "String", "Y", "결제영수증ID", "PAY-555"), ("newBalance", "Number", "Y", "남은 미수금", "0")]
            }
        ]
    },
    {
        "domain": "4_CRM",
        "apis": [
            {
                "name": "고객 리스트 조회", "method": "GET", "uri": "/api/v1/crm/guests", "desc": "CRM 고객 목록",
                "request": [("keyword", "Query", "N", "검색", "김철수")],
                "response": [("items", "Array", "Y", "배열", "[]"), ("items[].guestId", "String", "Y", "고객ID", "GST-01")]
            },
            {
                "name": "고객 정보 수정", "method": "PUT", "uri": "/api/v1/crm/guests/{id}", "desc": "연락처, 특별요청사항 등 고객 프로필 업데이트",
                "request": [("id", "Path", "Y", "고객ID", "GST-01"), ("phone", "Body", "N", "수정할 연락처", "010-0000-0000"), ("preferences", "Body", "N", "선호사항", "금연객실")],
                "response": [("success", "Boolean", "Y", "성공", "true")]
            }
        ]
    },
    {
        "domain": "5_Settings",
        "apis": [
            {
                "name": "직원 계정 조회", "method": "GET", "uri": "/api/v1/settings/staff", "desc": "직원 계정 리스트",
                "request": [("role", "Query", "N", "권한 필터", "admin")],
                "response": [("items", "Array", "Y", "배열", "[]"), ("items[].userId", "String", "Y", "계정ID", "USR-01")]
            },
            {
                "name": "신규 직원 등록", "method": "POST", "uri": "/api/v1/settings/staff", "desc": "새로운 관리자/직원 계정 생성",
                "request": [("username", "Body", "Y", "접속ID", "front_02"), ("password", "Body", "Y", "초기비밀번호", "Temp123!"), ("name", "Body", "Y", "실명", "이직원"), ("roleId", "Body", "Y", "부여할 권한ID", "ROL-02")],
                "response": [("success", "Boolean", "Y", "성공", "true"), ("userId", "String", "Y", "생성된ID", "USR-02")]
            },
            {
                "name": "호텔 마스터 설정 수정", "method": "PUT", "uri": "/api/v1/settings/general", "desc": "운영 정책(체크인 시간 등) 일괄 수정",
                "request": [("checkInTime", "Body", "Y", "체크인 기준", "15:00"), ("checkOutTime", "Body", "Y", "체크아웃 기준", "11:00")],
                "response": [("success", "Boolean", "Y", "성공", "true")]
            }
        ]
    }
]

# Create Excel
wb = Workbook()
wb.remove(wb.active)

header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
sub_header_fill = PatternFill(start_color="EAEFF7", end_color="EAEFF7", fill_type="solid")
sub_header_font = Font(bold=True)
border = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'), top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

for domain_data in api_data:
    ws = wb.create_sheet(title=domain_data['domain'])
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 25
    
    row_idx = 1
    for api in domain_data['apis']:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        cell = ws.cell(row=row_idx, column=1, value=f"■ [{api['method']}] {api['name']}")
        cell.font = Font(bold=True, size=11, color="1F4E78")
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1
        
        ws.cell(row=row_idx, column=1, value="URI").font = sub_header_font
        ws.cell(row=row_idx, column=1).fill = sub_header_fill
        ws.cell(row=row_idx, column=1).border = border
        ws.cell(row=row_idx, column=1).alignment = align_center
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
        c = ws.cell(row=row_idx, column=2, value=api['uri'])
        c.border = border
        c.font = Font(bold=True, color="C00000")
        for col in range(2, 7): ws.cell(row=row_idx, column=col).border = border
        row_idx += 1
        
        ws.cell(row=row_idx, column=1, value="설명").font = sub_header_font
        ws.cell(row=row_idx, column=1).fill = sub_header_fill
        ws.cell(row=row_idx, column=1).border = border
        ws.cell(row=row_idx, column=1).alignment = align_center
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
        c = ws.cell(row=row_idx, column=2, value=api['desc'])
        c.border = border
        for col in range(2, 7): ws.cell(row=row_idx, column=col).border = border
        row_idx += 1
        
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        req_title = ws.cell(row=row_idx, column=1, value="▶ Request (Input)")
        req_title.font = header_font
        req_title.fill = header_fill
        req_title.alignment = align_center
        row_idx += 1
        
        headers = ["구분 (I/O)", "파라미터/바디 필드", "데이터 타입", "필수 (Y/N)", "설명 (Description)", "샘플 (Sample)"]
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=h)
            c.font = sub_header_font
            c.fill = sub_header_fill
            c.border = border
            c.alignment = align_center
        row_idx += 1
        
        if not api['request']:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
            c = ws.cell(row=row_idx, column=1, value="요청 파라미터 없음")
            c.alignment = align_center
            c.border = border
            for col in range(1, 7): ws.cell(row=row_idx, column=col).border = border
            row_idx += 1
        else:
            for req in api['request']:
                row_data = ["Request", req[0], req[1], req[2], req[3], req[4]]
                for col_idx, val in enumerate(row_data, 1):
                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.border = border
                    c.alignment = align_left if col_idx in [2, 5] else align_center
                row_idx += 1
                
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        res_title = ws.cell(row=row_idx, column=1, value="▶ Response (Output)")
        res_title.font = header_font
        res_title.fill = header_fill
        res_title.alignment = align_center
        row_idx += 1
        
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=h)
            c.font = sub_header_font
            c.fill = sub_header_fill
            c.border = border
            c.alignment = align_center
        row_idx += 1
        
        for res in api['response']:
            row_data = ["Response", res[0], res[1], res[2], res[3], res[4]]
            for col_idx, val in enumerate(row_data, 1):
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.border = border
                c.alignment = align_left if col_idx in [2, 5] else align_center
            row_idx += 1
            
        row_idx += 2 

file_path = r'E:\AI_Project\Hotel_PMS\docs\API_Integration_Plan_Final_v5.xlsx'
wb.save(file_path)
print(f"Final Complete API Specification generated at: {file_path}")
