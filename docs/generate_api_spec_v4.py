import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

api_data = [
    {
        "domain": "1_Dashboard",
        "apis": [
            {
                "name": "대시보드 KPI (핵심지표) 조회", "method": "GET", "uri": "/api/v1/dashboard/kpis", "desc": "일일 점유율, ADR, RevPAR, 총매출 조회",
                "request": [("date", "Query", "Y", "기준 날짜", "2026-05-27")],
                "response": [("occupancyRate", "Number", "Y", "점유율", "85.4"), ("adr", "Number", "Y", "객실단가", "145.5"), ("revpar", "Number", "Y", "객실당수익", "124.2"), ("revenue", "Number", "Y", "매출", "24500")]
            },
            {
                "name": "매출 추이 차트", "method": "GET", "uri": "/api/v1/dashboard/revenue-chart", "desc": "날짜별 매출 데이터 배열 (차트용)",
                "request": [("startDate", "Query", "Y", "시작일", "2026-05-20"), ("endDate", "Query", "Y", "종료일", "2026-05-27")],
                "response": [("items", "Array", "Y", "데이터 배열", "[]"), ("items[].date", "String", "Y", "날짜", "2026-05-20"), ("items[].roomRev", "Number", "Y", "객실매출", "18500"), ("items[].fnbRev", "Number", "Y", "F&B매출", "3200")]
            }
        ]
    },
    {
        "domain": "2_Frontdesk",
        "apis": [
            {
                "name": "체크인/아웃 리스트 조회", "method": "GET", "uri": "/api/v1/frontdesk/checkins", "desc": "당일/지정일 체크인, 체크아웃 대상자 리스트",
                "request": [("date", "Query", "Y", "날짜", "2026-05-27"), ("type", "Query", "Y", "목록 타입 (arrival/departure)", "arrival"), ("keyword", "Query", "N", "이름/예약번호 검색", "홍길동")],
                "response": [("items", "Array", "Y", "목록", "[]"), ("items[].reservationId", "String", "Y", "예약ID", "RES-1001"), ("items[].guestName", "String", "Y", "투숙객명", "홍길동"), ("items[].roomTypeCode", "String", "Y", "객실타입", "DLX"), ("items[].roomNumber", "String", "N", "배정호실", "1401"), ("items[].balance", "Number", "Y", "잔액", "350")]
            },
            {
                "name": "예약 타임라인 (테이프차트) 조회", "method": "GET", "uri": "/api/v1/frontdesk/timeline", "desc": "객실별 날짜에 매핑된 예약 블록 데이터 조회",
                "request": [("startDate", "Query", "Y", "뷰 시작일", "2026-05-27"), ("days", "Query", "Y", "표시 일수", "14")],
                "response": [("rooms", "Array", "Y", "객실 배열", "[]"), ("rooms[].roomId", "String", "Y", "호실", "1401"), ("rooms[].reservations", "Array", "Y", "해당 호실의 예약 배열", "[...]"), ("rooms[].reservations[].resId", "String", "Y", "예약ID", "RES-101"), ("rooms[].reservations[].start", "String", "Y", "체크인", "2026-05-27"), ("rooms[].reservations[].end", "String", "Y", "체크아웃", "2026-05-29")]
            },
            {
                "name": "단체 예약 (Groups/Blocks) 조회", "method": "GET", "uri": "/api/v1/frontdesk/groups", "desc": "생성된 단체 블록 목록 조회",
                "request": [("status", "Query", "N", "상태 (tentative/definite)", "definite")],
                "response": [("items", "Array", "Y", "목록", "[]"), ("items[].blockId", "String", "Y", "블록ID", "GRP-01"), ("items[].name", "String", "Y", "행사명", "IT Summit"), ("items[].blockedRooms", "Number", "Y", "할당 객실", "50"), ("items[].pickedRooms", "Number", "Y", "예약 완료", "35")]
            },
            {
                "name": "고객사 (Companies/Agents) 조회", "method": "GET", "uri": "/api/v1/frontdesk/companies", "desc": "제휴 고객사 및 여행사 목록 조회",
                "request": [("type", "Query", "N", "타입 (corporate/travel_agent)", "corporate")],
                "response": [("items", "Array", "Y", "목록", "[]"), ("items[].companyId", "String", "Y", "고객사ID", "COM-01"), ("items[].name", "String", "Y", "회사명", "삼성전자"), ("items[].discountRate", "Number", "Y", "할인율", "15.0")]
            }
        ]
    },
    {
        "domain": "3_Operations",
        "apis": [
            {
                "name": "하우스키핑 객실 상태 조회", "method": "GET", "uri": "/api/v1/operations/housekeeping/rooms", "desc": "청소 및 점검용 객실 상태 그리드",
                "request": [("floor", "Query", "N", "층수", "14"), ("status", "Query", "N", "상태(clean/dirty)", "dirty")],
                "response": [("items", "Array", "Y", "목록", "[]"), ("items[].room", "String", "Y", "호실", "1401"), ("items[].hkStatus", "String", "Y", "청소상태", "dirty"), ("items[].frontStatus", "String", "Y", "점유상태", "occupied"), ("items[].guestFlag", "String", "N", "요청", "dnd")]
            },
            {
                "name": "유지보수 (Maintenance / OOS) 리스트", "method": "GET", "uri": "/api/v1/operations/maintenance", "desc": "고장(OOS) 및 수리 대기 객실/작업 리스트",
                "request": [("status", "Query", "N", "상태(pending/resolved)", "pending")],
                "response": [("items", "Array", "Y", "목록", "[]"), ("items[].issueId", "String", "Y", "이슈ID", "MNT-01"), ("items[].room", "String", "Y", "호실", "1401"), ("items[].type", "String", "Y", "분류(oos/ooh)", "oos"), ("items[].description", "String", "Y", "사유", "에어컨 고장")]
            },
            {
                "name": "통합 POS (부가서비스) 리스트", "method": "GET", "uri": "/api/v1/operations/pos", "desc": "룸서비스, 미니바, 세탁 등 오더 리스트",
                "request": [("category", "Query", "N", "분류", "roomservice")],
                "response": [("items", "Array", "Y", "목록", "[]"), ("items[].orderId", "String", "Y", "주문ID", "ORD-01"), ("items[].room", "String", "Y", "호실", "1401"), ("items[].category", "String", "Y", "분류", "roomservice"), ("items[].amount", "Number", "Y", "금액", "35.00")]
            },
            {
                "name": "객실 설정 (Room Setup) 리스트", "method": "GET", "uri": "/api/v1/operations/room-setup", "desc": "객실 번호, 타입, 층별 기본 설정 마스터 데이터",
                "request": [],
                "response": [("items", "Array", "Y", "목록", "[]"), ("items[].room", "String", "Y", "호실", "1401"), ("items[].type", "String", "Y", "객실타입", "DLX"), ("items[].features", "Array", "N", "특징", "['ocean_view']")]
            },
            {
                "name": "객실 요금제 (Rates) 마스터", "method": "GET", "uri": "/api/v1/operations/rates", "desc": "시즌별, 요일별 기준 단가 리스트",
                "request": [],
                "response": [("items", "Array", "Y", "목록", "[]"), ("items[].rateCode", "String", "Y", "요금제코드", "BAR"), ("items[].name", "String", "Y", "요금제명", "Best Available Rate"), ("items[].basePrice", "Number", "Y", "기본요금", "150.00")]
            },
            {
                "name": "나이트 오딧 (Night Audit) 상태 조회", "method": "GET", "uri": "/api/v1/operations/night-audit/status", "desc": "현재 호텔 시스템 날짜 및 일일 마감 진행 상태",
                "request": [],
                "response": [("businessDate", "String", "Y", "시스템 영업일", "2026-05-27"), ("status", "String", "Y", "마감 상태 (open/closed)", "open"), ("pendingCheckins", "Number", "Y", "미처리 체크인", "2"), ("pendingCheckouts", "Number", "Y", "미처리 체크아웃", "0")]
            },
            {
                "name": "통계/보고서 (Reports) 생성 요청", "method": "POST", "uri": "/api/v1/operations/reports", "desc": "매출, 점유율 등 각종 리포트 데이터 생성",
                "request": [("reportType", "Body", "Y", "리포트 종류 (revenue/occupancy)", "revenue"), ("startDate", "Body", "Y", "시작일", "2026-05-01"), ("endDate", "Body", "Y", "종료일", "2026-05-31")],
                "response": [("reportUrl", "String", "Y", "생성된 리포트 다운로드 URL", "https://.../report.pdf"), ("summaryData", "Object", "Y", "요약 데이터 객체", "{...}")]
            },
            {
                "name": "개별 정산/청구 (Folio) 조회", "method": "GET", "uri": "/api/v1/operations/folio/{resId}", "desc": "특정 예약의 전체 청구 내역(객실, F&B, 기타) 조회",
                "request": [("resId", "Path", "Y", "예약ID", "RES-1001")],
                "response": [("items", "Array", "Y", "청구 항목 배열", "[]"), ("items[].date", "String", "Y", "발생일", "2026-05-27"), ("items[].code", "String", "Y", "항목코드(ROOM, FNB)", "ROOM"), ("items[].amount", "Number", "Y", "청구액", "150.00"), ("totalDue", "Number", "Y", "총 미수금", "350.00")]
            }
        ]
    },
    {
        "domain": "4_CRM",
        "apis": [
            {
                "name": "고객 프로필 (Guests) 리스트", "method": "GET", "uri": "/api/v1/crm/guests", "desc": "등록된 고객 정보 및 히스토리 리스트",
                "request": [("keyword", "Query", "N", "이름/연락처", "김철수")],
                "response": [("items", "Array", "Y", "목록", "[]"), ("items[].guestId", "String", "Y", "고객ID", "GST-01"), ("items[].name", "String", "Y", "이름", "김철수"), ("items[].phone", "String", "N", "연락처", "010-1234"), ("items[].tier", "String", "Y", "멤버십", "GOLD"), ("items[].totalSpend", "Number", "Y", "총 결제액", "1500")]
            },
            {
                "name": "멤버십 (Membership) 등급 규정 조회", "method": "GET", "uri": "/api/v1/crm/membership", "desc": "호텔 멤버십 등급별 혜택 및 조건 마스터",
                "request": [],
                "response": [("items", "Array", "Y", "목록", "[]"), ("items[].tier", "String", "Y", "등급", "GOLD"), ("items[].requiredNights", "Number", "Y", "승급 필요 숙박수", "20"), ("items[].discount", "Number", "Y", "기본 할인율", "10.0")]
            }
        ]
    },
    {
        "domain": "5_Settings",
        "apis": [
            {
                "name": "직원 계정 (Staff) 리스트", "method": "GET", "uri": "/api/v1/settings/staff", "desc": "시스템에 등록된 사용자/직원 리스트",
                "request": [("role", "Query", "N", "권한 필터", "frontdesk")],
                "response": [("items", "Array", "Y", "목록", "[]"), ("items[].userId", "String", "Y", "유저ID", "USR-01"), ("items[].username", "String", "Y", "계정명", "admin01"), ("items[].name", "String", "Y", "이름", "김매니저"), ("items[].role", "String", "Y", "권한", "admin")]
            },
            {
                "name": "권한 (Roles) 마스터", "method": "GET", "uri": "/api/v1/settings/roles", "desc": "역할군별 시스템 접근 권한 마스터",
                "request": [],
                "response": [("items", "Array", "Y", "목록", "[]"), ("items[].roleId", "String", "Y", "권한ID", "ROL-01"), ("items[].name", "String", "Y", "권한명", "프론트데스크"), ("items[].permissions", "Array", "Y", "허용메뉴배열", "['checkin', 'folio']")]
            },
            {
                "name": "시스템 기본 설정 (General Settings)", "method": "GET", "uri": "/api/v1/settings/general", "desc": "호텔 기본 정보 (이름, 체크인시간, 통화 등)",
                "request": [],
                "response": [("propertyName", "String", "Y", "호텔명", "그랜드 호텔"), ("currency", "String", "Y", "기준 통화", "KRW"), ("checkInTime", "String", "Y", "체크인 기준시간", "15:00"), ("checkOutTime", "String", "Y", "체크아웃 기준시간", "11:00")]
            }
        ]
    }
]

# Create Excel
wb = Workbook()
wb.remove(wb.active)

header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
sub_header_fill = PatternFill(start_color="EAEFF7", end_color="EAEFF7", fill_type="solid")
sub_header_font = Font(bold=True)
border = Border(left=Side(style='thin', color='BFBFBF'), 
                right=Side(style='thin', color='BFBFBF'), 
                top=Side(style='thin', color='BFBFBF'), 
                bottom=Side(style='thin', color='BFBFBF'))
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

for domain_data in api_data:
    ws = wb.create_sheet(title=domain_data['domain'])
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 45
    ws.column_dimensions['F'].width = 25
    
    row_idx = 1
    for api in domain_data['apis']:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        cell = ws.cell(row=row_idx, column=1, value=f"■ [{api['method']}] {api['name']}")
        cell.font = Font(bold=True, size=11, color="1F4E78")
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1
        
        # URI
        ws.cell(row=row_idx, column=1, value="URI").font = sub_header_font
        ws.cell(row=row_idx, column=1).fill = sub_header_fill
        ws.cell(row=row_idx, column=1).border = border
        ws.cell(row=row_idx, column=1).alignment = align_center
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
        c = ws.cell(row=row_idx, column=2, value=api['uri'])
        c.border = border
        c.font = Font(bold=True, color="C00000")
        for col in range(2, 7): ws.cell(row=row_idx, column=col).border = border
        row_idx += 1
        
        # Description
        ws.cell(row=row_idx, column=1, value="설명").font = sub_header_font
        ws.cell(row=row_idx, column=1).fill = sub_header_fill
        ws.cell(row=row_idx, column=1).border = border
        ws.cell(row=row_idx, column=1).alignment = align_center
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
        c = ws.cell(row=row_idx, column=2, value=api['desc'])
        c.border = border
        for col in range(2, 7): ws.cell(row=row_idx, column=col).border = border
        row_idx += 1
        
        # Request Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        req_title = ws.cell(row=row_idx, column=1, value="▶ Request (Input)")
        req_title.font = header_font
        req_title.fill = header_fill
        req_title.alignment = align_center
        row_idx += 1
        
        headers = ["구분 (I/O)", "파라미터/바디 필드", "데이터 타입", "필수 (Y/N)", "설명 (Description)", "샘플 (Sample)"]
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=h)
            c.font = sub_header_font
            c.fill = sub_header_fill
            c.border = border
            c.alignment = align_center
        row_idx += 1
        
        if not api['request']:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
            c = ws.cell(row=row_idx, column=1, value="요청 파라미터 없음")
            c.alignment = align_center
            c.border = border
            for col in range(1, 7): ws.cell(row=row_idx, column=col).border = border
            row_idx += 1
        else:
            for req in api['request']:
                row_data = ["Request", req[0], req[1], req[2], req[3], req[4]]
                for col_idx, val in enumerate(row_data, 1):
                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.border = border
                    c.alignment = align_left if col_idx in [2, 5] else align_center
                row_idx += 1
                
        # Response Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        res_title = ws.cell(row=row_idx, column=1, value="▶ Response (Output)")
        res_title.font = header_font
        res_title.fill = header_fill
        res_title.alignment = align_center
        row_idx += 1
        
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=h)
            c.font = sub_header_font
            c.fill = sub_header_fill
            c.border = border
            c.alignment = align_center
        row_idx += 1
        
        for res in api['response']:
            row_data = ["Response", res[0], res[1], res[2], res[3], res[4]]
            for col_idx, val in enumerate(row_data, 1):
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.border = border
                c.alignment = align_left if col_idx in [2, 5] else align_center
            row_idx += 1
            
        row_idx += 2 # space between APIs

file_path = r'E:\AI_Project\Hotel_PMS\docs\API_Integration_Plan_Full.xlsx'
wb.save(file_path)
print(f"Full Menu API Specification generated at: {file_path}")
